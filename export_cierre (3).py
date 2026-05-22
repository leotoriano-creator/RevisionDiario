"""
export_cierre.py  v3.0
======================
Cambios v3.0:
    - HISTORIA: columna Hora agregada
    - HISTORIA: columnas con nombres reales por instrumento (TEM/TNA/TIREA/MD/WAL etc.)
    - Fix _d(): soporta formato DD/MM/YYYY (BONCER fecha vencimiento)
    - Backup diario sube a Drive en subcarpeta diarios/ via /api/admin/upload-cierre-diario
    - MD siempre con 2 decimales
"""

import os, sys, shutil, argparse, json
from datetime import datetime, date, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import google.auth.transport.requests as google_requests
    from google.oauth2 import service_account
    GOOGLE_AUTH_AVAILABLE = True
except ImportError:
    GOOGLE_AUTH_AVAILABLE = False

SERVICE_ACCOUNT_FILE = Path(__file__).parent / "service_account.json" 

# ─── Config ───────────────────────────────────────────────────────────────────
TZ_AR      = ZoneInfo("America/Argentina/Buenos_Aires")
API_BASE   = os.getenv("API_BASE_URL", "https://api.alquimiaconsultora.com").rstrip("/")
OUTPUT_DIR = Path(os.getenv("OUTPUT_DIR", "./output/reportes"))
EXPORT_META_JSON = Path(os.getenv("OUTPUT_DIR", "./output/reportes")).parent / "export_meta.json"

# ─── Paleta ───────────────────────────────────────────────────────────────────
BG_TITLE = "0A1520"; BG_HDR = "1F3864"; BG_ROW = "0D1B2A"; BG_ALT = "111E2D"
FG_WHITE = "FFFFFF"; FG_DATA = "D8E4F0"; FG_MUTED = "7A9BBB"
FG_TITLE = "00AEEF"; FG_BORD = "1E3355"

TAB_LECAPS = "00AEEF"; TAB_BONCER = "9B59B6"; TAB_DLK    = "E67E22"
TAB_TAMAR  = "27AE60"; TAB_DUAL   = "F39C12"; TAB_ONLN   = "4CAF50"
TAB_ONLNY  = "FF9800"; TAB_HDSOB  = "E74C3C"; TAB_BOTE   = "16A085"
TAB_HIST   = "888888"; TAB_DOLAR  = "2ECC71"; TAB_ADR = "3498DB"; TAB_CEDEAR = "E91E8C"; TAB_NYSE = "F39C12"; TAB_VARMON = "5DADE2"; TAB_ACCIONES = "A93226"
TAB_IDX_US   = "C0392B"; TAB_IDX_AM   = "8E44AD"; TAB_IDX_EU   = "2980B9"; TAB_IDX_ASIA = "16A085"
TAB_MONEDAS  = "D35400"; TAB_CMDTY    = "7F8C8D"; TAB_BONDS    = "1ABC9C"; TAB_CRIPTO   = "F1C40F"
TAB_ROFEX    = "27AE60"; TAB_CAUCIONES = "3498DB"

# ─── Estilo ───────────────────────────────────────────────────────────────────
def _fill(c): return PatternFill("solid", start_color=c, fgColor=c)
def _font(c=FG_DATA, bold=False, size=9): return Font(name="Arial", color=c, bold=bold, size=size)
def _center(): return Alignment(horizontal="center", vertical="center")
def _right():  return Alignment(horizontal="right",  vertical="center")
def _left():   return Alignment(horizontal="left",   vertical="center")
def _border():
    s = Side(style="thin", color=FG_BORD)
    return Border(left=s, right=s, top=s, bottom=s)

def paint_bg(ws, rows=300, cols=40):
    for r in range(1, rows+1):
        for c in range(1, cols+1):
            ws.cell(r, c).fill = _fill(BG_TITLE)

def write_meta(ws, titulo, fecha_dt, hora_dt, liq_dt):
    paint_bg(ws)
    ws.sheet_view.showGridLines = False
    ws["A1"].value = titulo
    ws["A1"].font  = Font(name="Arial", bold=True, color=FG_TITLE, size=13)
    ws["A1"].fill  = _fill(BG_TITLE)
    ws.row_dimensions[1].height = 22

    ws["A2"].value = "Precios al:"; ws["A2"].font = _font(FG_MUTED, size=8); ws["A2"].fill = _fill(BG_TITLE)
    ws["B2"].value = fecha_dt;      ws["B2"].number_format = "DD/MM/YYYY"
    ws["B2"].font  = _font(FG_DATA, bold=True, size=8); ws["B2"].fill = _fill(BG_TITLE)
    ws["C2"].value = hora_dt;       ws["C2"].number_format = "HH:MM"
    ws["C2"].font  = _font(FG_DATA, size=8); ws["C2"].fill = _fill(BG_TITLE)
    ws["D2"].value = "ART"; ws["D2"].font = _font(FG_MUTED, size=8); ws["D2"].fill = _fill(BG_TITLE)

    ws["A3"].value = "Liquidación:"; ws["A3"].font = _font(FG_MUTED, size=8); ws["A3"].fill = _fill(BG_TITLE)
    if liq_dt:
        ws["B3"].value = liq_dt; ws["B3"].number_format = "DD/MM/YYYY"
        ws["B3"].font  = _font(FG_DATA, size=8); ws["B3"].fill = _fill(BG_TITLE)

    ws.column_dimensions["T"].width = 14
    c = ws.cell(1, 20); c.value = "ALQUIMIA"
    c.font = Font(name="Arial", bold=True, color=FG_TITLE, size=11)
    c.fill = _fill(BG_TITLE); c.alignment = Alignment(horizontal="right", vertical="center")
    c2 = ws.cell(2, 20); c2.value = "Consultora"
    c2.font = _font(FG_MUTED, size=8); c2.fill = _fill(BG_TITLE)
    c2.alignment = Alignment(horizontal="right", vertical="center")

def write_hdr(ws, row, cols):
    ws.row_dimensions[row].height = 16
    for i, col in enumerate(cols, 1):
        c = ws.cell(row, i)
        c.value = col["label"]; c.fill = _fill(BG_HDR)
        c.font = _font(FG_WHITE, bold=True); c.alignment = _center(); c.border = _border()
        ws.column_dimensions[get_column_letter(i)].width = col.get("w", 11)

def write_row(ws, row, cols, data, alt=False):
    ws.row_dimensions[row].height = 15
    bg = BG_ALT if alt else BG_ROW
    for i, col in enumerate(cols, 1):
        c = ws.cell(row, i)
        c.value = data.get(col["key"]); c.fill = _fill(bg); c.font = _font(FG_DATA)
        c.alignment = _right() if col.get("align") == "right" else _center()
        c.border = _border()
        if col.get("fmt"): c.number_format = col["fmt"]

# ─── Utilidades ───────────────────────────────────────────────────────────────
def _f(v):
    try: return float(v) if v is not None else None
    except: return None

def _d(v):
    """Parsea fecha. Soporta YYYY-MM-DD y DD/MM/YYYY."""
    if v is None: return None
    s = str(v).strip()[:10]
    if len(s) == 10 and s[2] == "/" and s[5] == "/":
        try: return datetime.strptime(s, "%d/%m/%Y").date()
        except: return None
    try: return date.fromisoformat(s)
    except: return None

def _pct(v):
    x = _f(v)
    return None if x is None else x / 100.0

def api_get(path):
    try:
        r = requests.get(f"{API_BASE}{path}", timeout=30); r.raise_for_status()
        d = r.json(); return d if isinstance(d, list) else []
    except Exception as e:
        print(f"  [ERROR] GET {path}: {e}"); return []

def api_get_obj(path):
    try:
        r = requests.get(f"{API_BASE}{path}", timeout=30); r.raise_for_status()
        return r.json()
    except Exception as e:
        print(f"  [ERROR] GET {path}: {e}"); return {}

def _filtro_comun(precio, volumen, tir, vto, hoy):
    if vto and vto <= hoy:   return False
    if not precio or precio <= 0:   return False
    if not volumen or volumen <= 0: return False
    if tir is not None and (tir < -0.5 or tir > 5.0): return False
    return True

# ─── Historia ─────────────────────────────────────────────────────────────────
HIST_MAP = {
    "LECAPS": dict(dim="dias",     precio="precio",    var="var",  vol="volumen", tem="tem",  tna="tna",       tirea="tirea",       extra=None,        extra_lbl="",           dim_fmt="#,##0",  precio_fmt="#,##0.000", extra_fmt=""),
    "BONCER": dict(dim="md",       precio="precio",    var="var",  vol="volumen", tem=None,   tna="tna",       tirea="tirea",       extra=None,        extra_lbl="",           dim_fmt="0.00",   precio_fmt="#,##0.00",  extra_fmt=""),
    "DLK":    dict(dim="md",       precio="precio_ars",var="var",  vol="volumen", tem=None,   tna="tna",       tirea="tirea",       extra=None,        extra_lbl="",           dim_fmt="0.00",   precio_fmt="#,##0.00",  extra_fmt=""),
    "TAMAR":  dict(dim="dias_vto", precio="precio",    var="var",  vol="volumen", tem=None,   tna="tna_tamar", tirea="tirea_tamar", extra=None,        extra_lbl="",           dim_fmt="#,##0",  precio_fmt="#,##0.00",  extra_fmt=""),
    "DUAL":   dict(dim="dias_vto", precio="precio",    var="var",  vol="volumen", tem=None,   tna=None,        tirea="tirea_tamar", extra="tirea_fija",extra_lbl="TIREA Fija", dim_fmt="#,##0",  precio_fmt="#,##0.00",  extra_fmt="0.00%"),
    "ONS_LN": dict(dim="md",       precio="precio",    var="var",  vol="volumen", tem=None,   tna=None,        tirea="tirea",       extra="tirea_ccl", extra_lbl="TIREA CCL",  dim_fmt="0.00",   precio_fmt="#,##0.0000",extra_fmt="0.00%"),
    "ONS_LNY":dict(dim="md",       precio="precio",    var="var",  vol="volumen", tem=None,   tna=None,        tirea="tirea",       extra="tirea_ccl", extra_lbl="TIREA CCL",  dim_fmt="0.00",   precio_fmt="#,##0.0000",extra_fmt="0.00%"),
    "HDSOB":  dict(dim="md",       precio="precio",    var="var",  vol="volumen", tem=None,   tna="tna",       tirea="tirea",       extra=None,        extra_lbl="",           dim_fmt="0.00",   precio_fmt="#,##0.00",  extra_fmt=""),
    "BOTE":   dict(dim="md",       precio="precio_sucio",var="var",vol="volumen", tem=None,   tna="tna",       tirea="tirea",       extra="wal",       extra_lbl="WAL",        dim_fmt="0.00",   precio_fmt="#,##0.00",  extra_fmt="0.00"),
}
HIST_FIXED_HDRS   = ["Precio", "Var%",  "TEM",  "TNA",  "TIREA", "Días/MD", "Volumen", "Extra"]
HIST_FIXED_WIDTHS = [12,        8,       10,     10,     10,       9,          14,        13]


def load_historia_from_xlsx(ultimo_path: Path) -> dict:
    hist = {}
    if not ultimo_path.exists(): return hist
    try:
        wb = load_workbook(str(ultimo_path), read_only=True, data_only=True)
        if "HISTORIA" not in wb.sheetnames:
            wb.close(); return hist
        ws = wb["HISTORIA"]

        hdr = None
        for row in ws.iter_rows(min_row=3, max_row=3, values_only=True):
            hdr = row; break
        has_hora = hdr and len(hdr) > 2 and str(hdr[2] or "").upper() == "HORA"

        current = {}
        for row in ws.iter_rows(min_row=4, values_only=True):
            instr = row[0]
            if not instr or str(instr) == "Instrumento": continue
            instr = str(instr).upper()
            fecha = str(row[1])[:10] if row[1] else None
            if not fecha: continue

            if has_hora:
                hora_raw = row[2]
                if hasattr(hora_raw, "strftime"):
                    hora = hora_raw.strftime("%H:%M")
                elif hora_raw is not None:
                    hora = str(hora_raw)[:5]
                else:
                    hora = ""
                ticker  = str(row[3] or "")
                v_precio= row[4]
                v_var   = row[5]
                v_tem   = row[6]
                v_tna   = row[7]
                v_tirea = row[8]
                v_dim   = row[9]
                v_vol   = row[10]
                v_extra = row[11] if len(row) > 11 else None
            else:
                hora    = ""
                ticker  = str(row[2] or "")
                v_precio= None
                v_var   = None
                v_tem   = row[3]
                v_tna   = row[4]
                v_tirea = row[5]
                v_dim   = row[6] if len(row) > 6 else None
                v_vol   = row[7] if len(row) > 7 else None
                v_extra = row[8] if len(row) > 8 else None

            m = HIST_MAP.get(instr)
            if not m: continue

            snap_row = {"ticker": ticker}
            if m["dim"]:    snap_row[m["dim"]]    = v_dim
            if m["precio"]: snap_row[m["precio"]] = v_precio
            if m["var"]:    snap_row[m["var"]]    = v_var
            if m["vol"]:    snap_row[m["vol"]]    = v_vol
            if m["tem"]:    snap_row[m["tem"]]    = v_tem
            if m["tna"]:    snap_row[m["tna"]]    = v_tna
            if m["tirea"]:  snap_row[m["tirea"]]  = v_tirea
            if m["extra"]:  snap_row[m["extra"]]  = v_extra

            key = (instr, fecha, hora)
            if key not in current: current[key] = []
            current[key].append(snap_row)

        for (instr, fecha, hora), snap_rows in current.items():
            if instr not in hist: hist[instr] = []
            hist[instr].append({"fecha": fecha, "hora": hora, "snap": snap_rows})

        for instr in hist:
            hist[instr] = sorted(hist[instr], key=lambda h: h.get("fecha", ""))[-365:]

        wb.close()
    except Exception as e:
        print(f"  [HIST] Error leyendo historia: {e}")
    return hist


def add_snapshot_to_hist(hist, instr, fecha, hora, snap_rows):
    if instr not in hist: hist[instr] = []
    hist[instr] = [h for h in hist[instr] if h.get("fecha") != fecha]
    hist[instr].append({"fecha": fecha, "hora": hora, "snap": snap_rows})
    hist[instr] = sorted(hist[instr], key=lambda h: h.get("fecha", ""))[-365:]


# ═══════════════════════════════════════════════════════════════════════════════
# 1. LECAPS
# ═══════════════════════════════════════════════════════════════════════════════
LECAPS_COLS = [
    {"key":"ticker",       "label":"Especie",        "w": 8},
    {"key":"vto",          "label":"Vto.",            "w":11, "fmt":"DD/MM/YYYY"},
    {"key":"dias",         "label":"Días",            "w": 7, "align":"right"},
    {"key":"precio",       "label":"Precio Dirty",    "w":12, "align":"right", "fmt":"#,##0.000"},
    {"key":"var",          "label":"Var %",           "w": 9, "align":"right", "fmt":"0.00%"},
    {"key":"tem",          "label":"TEM (%)",         "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"tna",          "label":"TNA (%)",         "w":10, "align":"right", "fmt":"0.0%"},
    {"key":"tirea",        "label":"TIREA (%)",       "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"volumen",      "label":"Volumen",         "w":18, "align":"right", "fmt":"#,##0"},
    {"key":"vf",           "label":"Valor Final",     "w":12, "align":"right", "fmt":"#,##0.000"},
    {"key":"pct_directo",  "label":"% Directo",       "w":11, "align":"right", "fmt":"0.00%"},
    {"key":"mep_implicito","label":"MEP Implícito",   "w":14, "align":"right", "fmt":"#,##0.00"},
]

def _parse_lecaps(r, liq):
    vto   = _d(r.get("vencimiento") or r.get("fechaVencimiento"))
    tirea = _f(r.get("tea_bruta") or r.get("tirea"))
    tem   = _f(r.get("tem_bruta") or r.get("tem"))
    tna   = _f(r.get("tna_simple_bruta") or r.get("tna"))
    rv    = r.get("pct_change") if r.get("pct_change") is not None else r.get("variacionDiaria")
    if rv is None:
        p = _f(r.get("precio")); p0 = _f(r.get("precio_anterior") or r.get("precioAnterior"))
        rv = ((p/p0)-1)*100 if (p and p0 and p0 != 0) else None
    dias = r.get("dias_al_vto")
    if dias is None and vto and liq: dias = (vto - liq).days
    return {"ticker": (r.get("ticker") or "").strip(), "vto": vto,
            "dias": int(dias) if dias is not None else None,
            "precio": _f(r.get("precio")), "var": _pct(rv), "tem": tem, "tna": tna,
            "tirea": tirea, "volumen": _f(r.get("v") or r.get("volumen")),
            "vf": _f(r.get("valor_final") or r.get("valorFinal") or r.get("vf"))}

def _es_valido_lecaps(r, hoy):
    return _filtro_comun(_f(r.get("precio")), _f(r.get("v") or r.get("volumen")),
                         _f(r.get("tea_bruta") or r.get("tirea")),
                         _d(r.get("vencimiento") or r.get("fechaVencimiento")), hoy)

def build_lecaps(wb, hist, fecha_hoy, hora_str, fecha_dt, hora_dt):
    raw = api_get("/api/lecaps"); hoy = _d(fecha_hoy)
    liq = hoy + timedelta(days=1) if hoy else None
    rows = [_parse_lecaps(r, liq) for r in raw if _es_valido_lecaps(r, hoy)]
    rows.sort(key=lambda r: r.get("dias") or 9999)
    print(f"  {len(rows)} válidos  ({len(raw)-len(rows)} excluidos)")
    if not rows: return

    # MEP AL30 24hs desde dolares.json para % Directo y MEP Implícito
    mep_hoy = None
    try:
        r_dol = requests.get(DOLAR_URL, timeout=30); r_dol.raise_for_status()
        mep_hoy = _f(r_dol.json().get("data", {}).get("mep", {}).get("al30", {}).get("24hs", {}).get("price"))
        print(f"  MEP AL30 24hs para LECAPS: {mep_hoy}")
    except Exception as e:
        print(f"  [WARN] No se pudo obtener MEP para LECAPS: {e}")

    for r in rows:
        vf  = r.get("vf")
        px  = r.get("precio")
        if vf and px and px > 0:
            r["pct_directo"]   = (vf / px) - 1.0
            r["mep_implicito"] = (mep_hoy * (vf / px)) if mep_hoy else None
        else:
            r["pct_directo"]   = None
            r["mep_implicito"] = None

    snap = [{"ticker":r["ticker"],"dias":r["dias"],"precio":r["precio"],"var":r["var"],"volumen":r["volumen"],"tem":r["tem"],"tna":r["tna"],"tirea":r["tirea"]} for r in rows]
    add_snapshot_to_hist(hist, "LECAPS", fecha_hoy, hora_str, snap)
    ws = wb.create_sheet("LECAPS"); ws.sheet_properties.tabColor = TAB_LECAPS
    write_meta(ws, "LECAPS / BONCAPS — Tasa Fija ARS", fecha_dt, hora_dt, liq)
    write_hdr(ws, 5, LECAPS_COLS)
    for i, r in enumerate(rows): write_row(ws, 6+i, LECAPS_COLS, r, alt=(i%2==1))
    ws.freeze_panes = "A6"

# ═══════════════════════════════════════════════════════════════════════════════
# 2. BONCER
# ═══════════════════════════════════════════════════════════════════════════════
BONCER_COLS = [
    {"key":"ticker",      "label":"Ticker",          "w": 9},
    {"key":"vto",         "label":"Vencimiento",     "w":13, "fmt":"DD/MM/YYYY"},
    {"key":"precio",      "label":"Precio",          "w":12, "align":"right", "fmt":"#,##0.00"},
    {"key":"var",         "label":"% Var Diaria",    "w":11, "align":"right", "fmt":"0.00%"},
    {"key":"tirea",       "label":"TIREA (%)",       "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"tna",         "label":"TNA (%)",         "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"infl_impl",   "label":"Inflación Implícita", "w":16, "align":"right", "fmt":"0.00%"},
    {"key":"md",          "label":"MD",              "w": 8, "align":"right", "fmt":"0.00"},
    {"key":"vt",          "label":"VT",              "w":12, "align":"right", "fmt":"#,##0.00"},
    {"key":"paridad",     "label":"Paridad",         "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"ai",          "label":"Int. Corrido",    "w":12, "align":"right", "fmt":"#,##0.00"},
    {"key":"volumen",     "label":"Volumen",         "w":18, "align":"right", "fmt":"#,##0"},
]

def _parse_boncer(r):
    return {"ticker": (r.get("ticker") or "").strip(),
            "vto":    _d(r.get("fecha") or r.get("fechaVencimiento") or r.get("vencimiento")),
            "precio": _f(r.get("precio_sucio") or r.get("precioSucio")),
            "var":    _pct(r.get("variacion_diaria") or r.get("variacionDiaria")),
            "tirea":  _f(r.get("tirea")), "tna": _f(r.get("tna")),
            "md":     _f(r.get("modified_duration") or r.get("modifiedDuration")),
            "vt":     _f(r.get("valorTecnico")), "paridad": _f(r.get("paridad")),
            "ai":     _f(r.get("interesesCorridos")), "volumen": _f(r.get("volumen"))}

def _es_valido_boncer(r, hoy):
    return _filtro_comun(_f(r.get("precio_sucio") or r.get("precioSucio")), _f(r.get("volumen")),
                         _f(r.get("tirea")),
                         _d(r.get("fecha") or r.get("fechaVencimiento") or r.get("vencimiento")), hoy)

def _build_lecaps_nominal_index(hoy):
    """
    Devuelve un dict { (year, month): tirea_ea } con la TIREA nominal (EA, fracción)
    de la LECAP más cercana en días para cada mes de vencimiento.
    Replica el EoM-matching del frontend de BONCER.
    """
    raw = api_get("/api/lecaps")
    liq = hoy + timedelta(days=1) if hoy else None
    index = {}  # (year, month) -> {"tirea": float, "dias": int}
    for r in raw:
        vto  = _d(r.get("vencimiento") or r.get("fechaVencimiento"))
        tirea = _f(r.get("tea_bruta") or r.get("tirea"))
        if not vto or tirea is None: continue
        if vto <= hoy: continue
        dias = (vto - liq).days if liq else None
        key  = (vto.year, vto.month)
        # guarda la LECAP con menor dias (más cercana al EoM) para ese mes
        if key not in index or (dias is not None and dias < index[key]["dias"]):
            index[key] = {"tirea": tirea, "dias": dias or 0}
    # Convertir a fracción anual (vienen como decimal ya: 0.25 = 25%)
    return {k: v["tirea"] for k, v in index.items()}


def _infl_implicita_mensual(nominal_ea, real_ea):
    """
    Fisher anualizado → mensualizado. Replica exactamente el frontend:
      anual = (1+nominal) / (1+real) - 1
      mensual = (1+anual)^(30/360) - 1
    Ambas tasas en fracción (0.25 = 25%).
    """
    if nominal_ea is None or real_ea is None: return None
    try:
        i_anual = (1 + nominal_ea) / (1 + real_ea) - 1
        return (1 + i_anual) ** (30 / 360) - 1
    except Exception:
        return None


def build_boncer(wb, hist, fecha_hoy, hora_str, fecha_dt, hora_dt):
    data = api_get_obj("/api/boncer")
    raw  = data.get("rows",[]) if isinstance(data,dict) else (data if isinstance(data,list) else [])
    hoy  = _d(fecha_hoy); liq = hoy + timedelta(days=1) if hoy else None
    rows = [_parse_boncer(r) for r in raw if _es_valido_boncer(r, hoy)]
    rows.sort(key=lambda r: r.get("vto") or date.max)
    print(f"  {len(rows)} válidos  ({len(raw)-len(rows)} excluidos)")
    if not rows: return

    # Índice de tasas nominales LECAPS por mes (para inflación implícita)
    print("  Cargando LECAPS para inflación implícita BONCER...")
    lecaps_idx = _build_lecaps_nominal_index(hoy)
    print(f"  LECAPS index: {len(lecaps_idx)} meses ({list(lecaps_idx.keys())[:4]}...)")

    for r in rows:
        vto = r.get("vto")
        real_ea = r.get("tirea")  # ya es fracción anual
        if vto is None or real_ea is None:
            r["infl_impl"] = None
            continue
        # Buscar nominal del mismo mes; si no hay, buscar ±2 meses (igual que el front)
        nominal_ea = None
        key = (vto.year, vto.month)
        if key in lecaps_idx:
            nominal_ea = lecaps_idx[key]
        else:
            best_dist = None
            for off in [-2, -1, 1, 2]:
                # calcular mes relativo
                m = vto.month - 1 + off  # 0-indexed
                y = vto.year + m // 12
                m = m % 12 + 1
                alt_key = (y, m)
                if alt_key in lecaps_idx:
                    if best_dist is None or abs(off) < best_dist:
                        best_dist = abs(off)
                        nominal_ea = lecaps_idx[alt_key]
        r["infl_impl"] = _infl_implicita_mensual(nominal_ea, real_ea)

    snap = [{"ticker":r["ticker"],"md":r["md"],"precio":r["precio"],"var":r["var"],"volumen":r["volumen"],"tirea":r["tirea"],"tna":r["tna"]} for r in rows]
    add_snapshot_to_hist(hist, "BONCER", fecha_hoy, hora_str, snap)
    ws = wb.create_sheet("BONCER"); ws.sheet_properties.tabColor = TAB_BONCER
    write_meta(ws, "BONCER — Bonos CER-Linked ARS", fecha_dt, hora_dt, liq)
    write_hdr(ws, 5, BONCER_COLS)
    for i, r in enumerate(rows): write_row(ws, 6+i, BONCER_COLS, r, alt=(i%2==1))
    ws.freeze_panes = "A6"

# ═══════════════════════════════════════════════════════════════════════════════
# 3. DLK
# ═══════════════════════════════════════════════════════════════════════════════
DLK_COLS = [
    {"key":"ticker",     "label":"Ticker",      "w": 9},
    {"key":"emisor",     "label":"Emisor",       "w":20},
    {"key":"vto",        "label":"Vencimiento",  "w":13, "fmt":"DD/MM/YYYY"},
    {"key":"precio_ars", "label":"Precio ARS",   "w":14, "align":"right", "fmt":"#,##0.00"},
    {"key":"precio_usd", "label":"Precio USD",   "w":12, "align":"right", "fmt":"#,##0.0000"},
    {"key":"var",        "label":"% Var Diaria", "w":11, "align":"right", "fmt":"0.00%"},
    {"key":"tirea",      "label":"TIREA (%)",    "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"tna",        "label":"TNA (%)",      "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"md",         "label":"MD",           "w": 8, "align":"right", "fmt":"0.00"},
    {"key":"vt",         "label":"VT",           "w": 8, "align":"right", "fmt":"#,##0.00"},
    {"key":"paridad",    "label":"Paridad",      "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"ai",         "label":"Int. Corrido", "w":12, "align":"right", "fmt":"#,##0.0000"},
    {"key":"volumen",    "label":"Volumen",      "w":15, "align":"right", "fmt":"#,##0"},
    {"key":"a3500",      "label":"A3500",        "w":10, "align":"right", "fmt":"#,##0.00"},
]

def _parse_dlk(r):
    return {"ticker": (r.get("ticker") or "").strip(), "emisor": r.get("emisor") or "",
            "vto": _d(r.get("fecha") or r.get("fechaVencimiento")),
            "precio_ars": _f(r.get("precio_ars")),
            "precio_usd": _f(r.get("precio_limpio") or r.get("precioUsd")),
            "var": _pct(r.get("variacion_diaria")),
            "tirea": _f(r.get("tirea")), "tna": _f(r.get("tna")),
            "md": _f(r.get("modified_duration")), "vt": _f(r.get("vt")),
            "paridad": _f(r.get("paridad")), "ai": _f(r.get("ai_usd")),
            "volumen": _f(r.get("volumen")), "a3500": _f(r.get("a3500"))}

def _es_valido_dlk(r, hoy):
    return _filtro_comun(_f(r.get("precio_ars")), _f(r.get("volumen")), _f(r.get("tirea")),
                         _d(r.get("fecha") or r.get("fechaVencimiento")), hoy)

def build_dlk(wb, hist, fecha_hoy, hora_str, fecha_dt, hora_dt):
    data = api_get_obj("/api/dlk")
    raw  = data.get("rows",[]) if isinstance(data,dict) else (data if isinstance(data,list) else [])
    hoy  = _d(fecha_hoy); liq = hoy + timedelta(days=1) if hoy else None
    rows = [_parse_dlk(r) for r in raw if _es_valido_dlk(r, hoy)]
    rows.sort(key=lambda r: r.get("vto") or date.max)
    print(f"  {len(rows)} válidos  ({len(raw)-len(rows)} excluidos)")
    if not rows: return
    snap = [{"ticker":r["ticker"],"md":r["md"],"precio_ars":r["precio_ars"],"var":r["var"],"volumen":r["volumen"],"tirea":r["tirea"],"tna":r["tna"]} for r in rows]
    add_snapshot_to_hist(hist, "DLK", fecha_hoy, hora_str, snap)
    ws = wb.create_sheet("DLK"); ws.sheet_properties.tabColor = TAB_DLK
    write_meta(ws, "Dólar Linked — Bonos atados al A3500", fecha_dt, hora_dt, liq)
    write_hdr(ws, 5, DLK_COLS)
    for i, r in enumerate(rows): write_row(ws, 6+i, DLK_COLS, r, alt=(i%2==1))
    ws.freeze_panes = "A6"


# ═══════════════════════════════════════════════════════════════════════════════
# SENSIBILIDAD TAMAR/DUALES
# ═══════════════════════════════════════════════════════════════════════════════

# Parámetros de la tabla de sensibilidad (igual que el front)
TAMAR_SENS_MIN  = 0.20   # 20%
TAMAR_SENS_MAX  = 0.40   # 40%
TAMAR_SENS_STEP = 0.025  # 2.5%

def _yearfrac_30_360_us(d1, d2):
    """Fracción de año 30/360 US (NASD). Replica yearfrac_30_360_US del front."""
    Y1, M1, D1 = d1.year, d1.month, d1.day
    Y2, M2, D2 = d2.year, d2.month, d2.day
    # Feb EOM
    import calendar
    def is_eom(d): return d.day == calendar.monthrange(d.year, d.month)[1]
    if M1 == 2 and is_eom(d1): D1 = 30
    if M2 == 2 and is_eom(d2) and D1 >= 30: D2 = 30
    if D1 == 31: D1 = 30
    if D2 == 31 and (D1 == 30 or D1 == 31): D2 = 30
    return (360*(Y2-Y1) + 30*(M2-M1) + (D2-D1)) / 360

def _tem_from_tna(tna_unit):
    """TEM desde TNA aplicable + margen. m = 365/32 períodos/año."""
    m = 365 / 32
    return (1 + tna_unit / m) ** (m / 12) - 1

def _compute_vf_tamar(fecha_emision, fecha_vto, tna_aplic_unit, spread_unit):
    """
    VF = 100 × (1+TEM)^(yearfrac_30_360_us(emis,vto) × 12)
    Replica computeVFTamar del front.
    """
    if fecha_emision is None or fecha_vto is None: return None
    yf = _yearfrac_30_360_us(fecha_emision, fecha_vto)
    if not yf or yf <= 0: return None
    tna_total = (tna_aplic_unit or 0) + (spread_unit or 0)
    if tna_total is None or tna_total != tna_total: return None  # nan check
    tem = _tem_from_tna(tna_total)
    return 100 * (1 + tem) ** (yf * 12)

def _tna_aplicable(tamar_actual, tamar_proy, fecha_emision, fecha_liq, fecha_vto, dias_rest):
    """
    Promedio ponderado por días: dev × diasDev + proy × diasRest / total.
    Replica tnaAplicableFromRow del front.
    """
    dev  = tamar_actual
    proj = tamar_proy
    if dev is None and proj is None: return None
    if dev is None: return proj
    if proj is None: return dev

    # días devengados (emision → liq)
    dias_dev = None
    if fecha_emision and fecha_liq:
        dias_dev = (fecha_liq - fecha_emision).days
    # días restantes
    if dias_rest is None and fecha_liq and fecha_vto:
        dias_rest = (fecha_vto - fecha_liq).days

    if dias_dev is not None and dias_rest is not None and (dias_dev + dias_rest) > 0:
        return (dev * dias_dev + proj * dias_rest) / (dias_dev + dias_rest)
    # fallback: promedio simple
    return (dev + proj) / 2

def _tirea_dado_tamar_proj(precio, dias, vf):
    """TIREA = (VF/Precio)^(365/dias) - 1"""
    if not precio or not dias or not vf or dias <= 0: return None
    try:
        return (vf / precio) ** (365 / dias) - 1
    except Exception:
        return None

def _breakeven_tamar(precio, dias, fecha_emision, fecha_vto, fecha_liq,
                     tamar_actual, spread_unit, ref_tirea,
                     lo=0.0, hi=3.0, iters=64):
    """
    Bisección: TAMAR proyectada tal que TIREA resultante = ref_tirea.
    Replica breakevenTamar del front.
    """
    if ref_tirea is None: return None

    def tirea_at(tamar_proj):
        aplic = _tna_aplicable(tamar_actual, tamar_proj, fecha_emision, fecha_liq, fecha_vto, dias)
        vf    = _compute_vf_tamar(fecha_emision, fecha_vto, aplic, spread_unit)
        return _tirea_dado_tamar_proj(precio, dias, vf)

    t_lo = tirea_at(lo)
    t_hi = tirea_at(hi)
    if t_lo is None or t_hi is None: return None
    if t_lo >= ref_tirea or t_hi <= ref_tirea: return None

    for _ in range(iters):
        mid   = (lo + hi) / 2
        t_mid = tirea_at(mid)
        if t_mid is None: break
        if t_mid < ref_tirea: lo = mid
        else:                  hi = mid
        if (hi - lo) < 1e-8: break
    return (lo + hi) / 2

def _build_sens_tamar_raw(rows_tamar, rows_duales, raw_lecaps, liq):
    """
    Construye la tabla de sensibilidad para ambos grupos.
    Devuelve lista de dicts con todos los campos para escribir en Excel.
    """
    import math

    # Índice LECAPS: dias_al_vto → tirea_ea (fracción) — para LECAP comparable
    lecap_index = []  # lista de (dias, tirea)
    for r in (raw_lecaps or []):
        dias = r.get("dias_al_vto") or r.get("diasAlVto")
        tirea = _f(r.get("tea_bruta") or r.get("tirea"))
        if dias is not None and tirea is not None:
            lecap_index.append((int(dias), tirea))

    def lecap_comparable(dias_target):
        if not lecap_index or dias_target is None: return None
        best = min(lecap_index, key=lambda x: abs(x[0] - dias_target))
        return best[1]

    # LECAP ref: promedio simple de todas las TIREA (= autoRefFija del front)
    ref_tirea = None
    if lecap_index:
        vals = [t for _, t in lecap_index if t and t > 0]
        if vals:
            avg = sum(vals) / len(vals)
            ref_tirea = avg if abs(avg) < 1 else avg / 100

    # rangos de TAMAR proyectada
    rates = []
    v = TAMAR_SENS_MIN
    while v <= TAMAR_SENS_MAX + 1e-9:
        rates.append(round(v, 8))
        v += TAMAR_SENS_STEP

    result = []

    def process_row(r, grupo):
        precio   = _f(r.get("precio") or r.get("precioSucio"))
        dias     = r.get("dias_vto") or r.get("diasAlVto") or r.get("dias_al_vto")
        if dias: dias = int(dias)
        vto      = r.get("vto") or _d(r.get("fechaVencimiento") or r.get("vencimiento"))
        emis_raw = r.get("fechaEmision") or r.get("emision") or r.get("issueDate")
        fecha_emis = _d(emis_raw) if isinstance(emis_raw, str) else emis_raw

        if precio is None or precio <= 0: return

        # spread
        spread_pct = _f(r.get("spreadTamarPct"))
        spread_bps = _f(r.get("spreadTamarBps") or r.get("spread_bps") or r.get("spread"))
        if spread_pct is not None:
            spread_unit = spread_pct
        elif spread_bps is not None:
            spread_unit = spread_bps / 10000
        else:
            spread_unit = 0.0

        # TAMAR actual y esperada
        tamar_actual = _f(r.get("tamarActualTNA") or r.get("tamar_actual_tna"))
        if tamar_actual and abs(tamar_actual) > 3: tamar_actual /= 100
        tamar_esp    = _f(r.get("tamarEsperadaTNA") or r.get("tamar_esperada_tna"))
        if tamar_esp and abs(tamar_esp) > 3: tamar_esp /= 100

        # VF y TIREA de mercado (usando TAMAR esperada)
        aplic_mkt = _tna_aplicable(tamar_actual, tamar_esp, fecha_emis, liq, vto, dias)
        vf_mkt    = _compute_vf_tamar(fecha_emis, vto, aplic_mkt, spread_unit)
        tirea_mkt = _tirea_dado_tamar_proj(precio, dias, vf_mkt)

        # LECAP comparable
        lec_comp = lecap_comparable(dias)

        # Breakeven
        be = _breakeven_tamar(precio, dias, fecha_emis, vto, liq,
                               tamar_actual, spread_unit, ref_tirea)

        # Escenarios
        escenarios = {}
        for tamar_proj in rates:
            aplic = _tna_aplicable(tamar_actual, tamar_proj, fecha_emis, liq, vto, dias)
            vf_sc = _compute_vf_tamar(fecha_emis, vto, aplic, spread_unit)
            tirea_sc = _tirea_dado_tamar_proj(precio, dias, vf_sc)
            escenarios[tamar_proj] = tirea_sc

        result.append({
            "grupo":     grupo,
            "ticker":    r.get("ticker", ""),
            "vto":       vto,
            "precio":    precio,
            "vf_mkt":    vf_mkt,
            "tirea_mkt": tirea_mkt,
            "breakeven": be,
            "lec_comp":  lec_comp,
            "ref_tirea": ref_tirea,
            "escenarios": escenarios,
            "rates":     rates,
        })

    for r in (rows_tamar  or []): process_row(r, "TAMAR")
    for r in (rows_duales or []): process_row(r, "DUALES")
    return result, rates, ref_tirea


def _write_sens_sheet(ws, sens_rows, rates, ref_tirea,
                      titulo, tab_color, fecha_dt, hora_dt):
    """Escribe la hoja de sensibilidad en el workbook."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils  import get_column_letter

    ws.sheet_properties.tabColor = tab_color
    ws.sheet_view.showGridLines  = False
    write_meta(ws, titulo, fecha_dt, hora_dt, None)

    nf2 = lambda v: f"{v*100:.2f}%" if v is not None else "—"

    # ── colores heatmap ───────────────────────────────────────────────────────
    def heat_fill(var_pct):
        """var_pct es la variación % sobre precio (−22 a +12). Retorna PatternFill."""
        if var_pct is None: return _fill(BG_ROW)
        clamp = max(-22, min(12, var_pct))
        t     = (clamp + 22) / (12 + 22)
        hue   = int(t * 120)  # 0=rojo, 120=verde
        # Convertir HSL aproximado a RGB
        import colorsys
        r, g, b = colorsys.hls_to_rgb(hue/360, 0.38, 0.72)
        hex_color = f"{int(r*255):02X}{int(g*255):02X}{int(b*255):02X}"
        return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

    # ── cabecera fija ─────────────────────────────────────────────────────────
    ref_lbl = "Lecap comp. / ref: " + nf2(ref_tirea)
    FIXED_LABELS = ["Especie", "Vto", "Precio", "VF (mkt)", "TIREA (mkt)",
                    "Breakeven", ref_lbl]
    FIXED_WIDTHS = [10, 13, 10, 10, 12, 12, 16]
    HDR_ROW = 5

    ws.row_dimensions[HDR_ROW].height = 28

    # Fila 1 del header: cols fijas + "TAMAR proy. — TIREA resultante" (merged)
    n_fixed = len(FIXED_LABELS)
    n_rates  = len(rates)

    for i, (lbl, w) in enumerate(zip(FIXED_LABELS, FIXED_WIDTHS), 1):
        c = ws.cell(HDR_ROW, i)
        c.value = lbl
        c.fill  = _fill(BG_HDR)
        c.font  = _font(FG_WHITE, bold=True, size=8)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()
        ws.column_dimensions[get_column_letter(i)].width = w

    # Header de los escenarios: una celda por tasa
    for j, rate in enumerate(rates):
        col = n_fixed + 1 + j
        c   = ws.cell(HDR_ROW, col)
        c.value = rate
        c.number_format = "0.00%"
        c.fill  = _fill(BG_HDR)
        c.font  = _font(FG_WHITE, bold=True, size=8)
        c.alignment = _center()
        c.border = _border()
        ws.column_dimensions[get_column_letter(col)].width = 9

    # ── filas de datos ────────────────────────────────────────────────────────
    row_idx = HDR_ROW + 1
    current_group = None

    for s in sens_rows:
        # separador de grupo
        if s["grupo"] != current_group:
            current_group = s["grupo"]
            label_sep = "TAMAR — Tasa variable" if current_group == "TAMAR"                         else "DUALES — Tasa variable + Fija"
            color_sep = TAB_TAMAR if current_group == "TAMAR" else "A78BFA"
            total_cols = n_fixed + n_rates
            for col in range(1, total_cols + 1):
                c = ws.cell(row_idx, col)
                c.fill   = _fill(color_sep + "30" if len(color_sep) == 6 else BG_ALT)
                c.border = _border()
                c.font   = Font(name="Arial", bold=True, color=color_sep if len(color_sep)==6 else FG_TITLE, size=8)
            ws.cell(row_idx, 1).value = label_sep
            ws.cell(row_idx, 1).font  = Font(name="Arial", bold=True,
                                             color=color_sep if len(color_sep)==6 else FG_TITLE,
                                             size=8)
            ws.row_dimensions[row_idx].height = 14
            row_idx += 1

        ws.row_dimensions[row_idx].height = 15
        bg = BG_ROW

        # cols fijas
        def cell(col, val, fmt=None, bold=False, color=FG_DATA):
            c = ws.cell(row_idx, col)
            c.value  = val
            c.fill   = _fill(bg)
            c.font   = Font(name="Arial", color=color, bold=bold, size=8)
            c.border = _border()
            c.alignment = _center()
            if fmt: c.number_format = fmt

        cell(1, s["ticker"],    bold=True)
        if s["vto"]: ws.cell(row_idx,2).value = s["vto"]; ws.cell(row_idx,2).number_format="DD/MM/YYYY"
        ws.cell(row_idx,2).fill=_fill(bg); ws.cell(row_idx,2).font=_font(FG_DATA,size=8); ws.cell(row_idx,2).border=_border(); ws.cell(row_idx,2).alignment=_center()
        cell(3, s["precio"],    "#,##0.00")
        cell(4, s["vf_mkt"],    "#,##0.00")
        cell(5, s["tirea_mkt"], "0.00%")
        # breakeven en amarillo si existe
        c_be = ws.cell(row_idx, 6)
        c_be.value  = s["breakeven"]
        c_be.fill   = _fill(bg)
        c_be.font   = Font(name="Arial", color="FBBF24" if s["breakeven"] else FG_MUTED, bold=True, size=8)
        c_be.border = _border(); c_be.alignment = _center()
        if s["breakeven"]: c_be.number_format = "0.00%"
        # lecap comparable en verde
        c_lc = ws.cell(row_idx, 7)
        c_lc.value  = s["lec_comp"]
        c_lc.fill   = _fill(bg)
        c_lc.font   = Font(name="Arial", color="34D399", size=8)
        c_lc.border = _border(); c_lc.alignment = _center()
        if s["lec_comp"]: c_lc.number_format = "0.00%"

        # escenarios con heatmap
        for j, rate in enumerate(rates):
            col     = n_fixed + 1 + j
            tirea_sc = s["escenarios"].get(rate)
            # var% = (precio_implicito/precio - 1)*100 para el color
            var_pct = None
            if tirea_sc is not None and s["ref_tirea"] is not None and s["vf_mkt"] is not None and s["vf_mkt"] > 0:
                dias = None
                if s["vto"] and s["vto"]:
                    from datetime import date as _date
                    # aproximar días desde hoy (solo para el color)
                    try:
                        import datetime
                        dias_approx = (s["vto"] - datetime.date.today()).days
                        if dias_approx > 0 and tirea_sc is not None and s["ref_tirea"] is not None:
                            p_impl = s["vf_mkt"] / ((1 + s["ref_tirea"]) ** (dias_approx / 365))
                            var_pct = (p_impl / s["precio"] - 1) * 100 if s["precio"] else None
                    except Exception:
                        pass
            c_sc = ws.cell(row_idx, col)
            c_sc.value  = tirea_sc
            c_sc.fill   = heat_fill(var_pct)
            c_sc.font   = Font(name="Arial", color="FFFFFF", bold=True, size=8)
            c_sc.border = _border()
            c_sc.alignment = _center()
            if tirea_sc is not None: c_sc.number_format = "0.00%"

        row_idx += 1

    ws.freeze_panes = f"A{HDR_ROW+1}"
    print(f"  {row_idx - HDR_ROW - 1} filas en sensibilidad")


# ═══════════════════════════════════════════════════════════════════════════════
# 4. TAMAR
# ═══════════════════════════════════════════════════════════════════════════════
TAMAR_COLS = [
    {"key":"ticker",      "label":"Ticker",            "w": 9},
    {"key":"vto",         "label":"Vencimiento",       "w":13, "fmt":"DD/MM/YYYY"},
    {"key":"dias_vto",    "label":"Días al Vto",       "w": 9, "align":"right"},
    {"key":"precio",      "label":"Precio",            "w":10, "align":"right", "fmt":"#,##0.00"},
    {"key":"var",         "label":"% Var Diaria",      "w":11, "align":"right", "fmt":"0.00%"},
    {"key":"tem_tamar",   "label":"TEM TAMAR",         "w":11, "align":"right", "fmt":"0.00%"},
    {"key":"tna_tamar",   "label":"TNA TAMAR",         "w":11, "align":"right", "fmt":"0.00%"},
    {"key":"tirea_tamar", "label":"TIREA TAMAR",       "w":11, "align":"right", "fmt":"0.00%"},
    {"key":"vf_tamar",    "label":"Valor Final",       "w":11, "align":"right", "fmt":"#,##0.00"},
    {"key":"tamar_aplic", "label":"TAMAR Aplic (TNA)", "w":15, "align":"right", "fmt":"0.00%"},
    {"key":"spread_bps",  "label":"Spread (bps)",      "w":12, "align":"right", "fmt":"#,##0.00"},
    {"key":"volumen",     "label":"Volumen",           "w":18, "align":"right", "fmt":"#,##0"},
]

def _parse_tamar(r):
    return {"ticker": (r.get("ticker") or "").strip(),
            "vto": _d(r.get("fechaVencimiento")), "dias_vto": r.get("diasAlVto"),
            "precio": _f(r.get("precioSucio")), "var": _pct(r.get("variacionDiaria")),
            "tem_tamar": _f(r.get("temTamar")), "tna_tamar": _f(r.get("tnaTamar")),
            "tirea_tamar": _f(r.get("tireaTamar")), "vf_tamar": _f(r.get("valorFinalTamar")),
            "tamar_aplic": _f(r.get("tamarTnaAplicable")), "spread_bps": _f(r.get("spreadTamarBps")),
            "volumen": _f(r.get("volumen"))}

def _es_valido_tamar(r, hoy):
    return _filtro_comun(_f(r.get("precioSucio")), _f(r.get("volumen")), _f(r.get("tireaTamar")),
                         _d(r.get("fechaVencimiento")), hoy)

def build_tamar(wb, hist, fecha_hoy, hora_str, fecha_dt, hora_dt):
    raw = api_get("/api/tamar"); hoy = _d(fecha_hoy); liq = hoy + timedelta(days=1) if hoy else None
    rows = [_parse_tamar(r) for r in raw if _es_valido_tamar(r, hoy)]
    rows.sort(key=lambda r: r.get("vto") or date.max)
    print(f"  {len(rows)} válidos  ({len(raw)-len(rows)} excluidos)")
    if not rows: return
    snap = [{"ticker":r["ticker"],"dias_vto":r["dias_vto"],"precio":r["precio"],"var":r["var"],"volumen":r["volumen"],"tna_tamar":r["tna_tamar"],"tirea_tamar":r["tirea_tamar"]} for r in rows]
    add_snapshot_to_hist(hist, "TAMAR", fecha_hoy, hora_str, snap)
    ws = wb.create_sheet("TAMAR"); ws.sheet_properties.tabColor = TAB_TAMAR
    write_meta(ws, "TAMAR — Bonos tasa variable TAMAR", fecha_dt, hora_dt, liq)
    write_hdr(ws, 5, TAMAR_COLS)
    for i, r in enumerate(rows): write_row(ws, 6+i, TAMAR_COLS, r, alt=(i%2==1))
    ws.freeze_panes = "A6"

# ═══════════════════════════════════════════════════════════════════════════════
# 5. DUALES
# ═══════════════════════════════════════════════════════════════════════════════
DUAL_COLS = [
    {"key":"ticker",      "label":"Ticker",            "w": 9},
    {"key":"vto",         "label":"Vencimiento",       "w":13, "fmt":"DD/MM/YYYY"},
    {"key":"precio",      "label":"Precio",            "w":10, "align":"right", "fmt":"#,##0.00"},
    {"key":"var",         "label":"% Var Diaria",      "w":11, "align":"right", "fmt":"0.00%"},
    {"key":"tem_tamar",   "label":"TEM TAMAR",         "w":11, "align":"right", "fmt":"0.00%"},
    {"key":"tna_tamar",   "label":"TNA TAMAR",         "w":11, "align":"right", "fmt":"0.00%"},
    {"key":"tirea_tamar", "label":"TIREA TAMAR",       "w":11, "align":"right", "fmt":"0.00%"},
    {"key":"vf_tamar",    "label":"VF TAMAR",          "w":10, "align":"right", "fmt":"#,##0.00"},
    {"key":"dias_vto",    "label":"Días al Vto",       "w": 9, "align":"right"},
    {"key":"tem_fija",    "label":"TEM Fija",          "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"tna_fija",    "label":"TNA Fija",          "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"tirea_fija",  "label":"TIREA Fija",        "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"vf_fija",     "label":"VF Fija",           "w":10, "align":"right", "fmt":"#,##0.00"},
    {"key":"tamar_aplic", "label":"TAMAR Aplic (TNA)", "w":15, "align":"right", "fmt":"0.00%"},
    {"key":"pata",        "label":"Pata Elegida",      "w":12},
    {"key":"spread_bps",  "label":"Spread (bps)",      "w":12, "align":"right", "fmt":"#,##0.00"},
    {"key":"volumen",     "label":"Volumen",           "w":18, "align":"right", "fmt":"#,##0"},
]

def _parse_dual(r):
    return {"ticker": (r.get("ticker") or "").strip(),
            "vto": _d(r.get("fechaVencimiento")), "dias_vto": r.get("diasAlVto"),
            "precio": _f(r.get("precioSucio")), "var": _pct(r.get("variacionDiaria")),
            "tem_tamar": _f(r.get("temTamar")), "tna_tamar": _f(r.get("tnaTamar")),
            "tirea_tamar": _f(r.get("tireaTamar")), "vf_tamar": _f(r.get("valorFinalTamar")),
            "tem_fija": _f(r.get("temFija")), "tna_fija": _f(r.get("tnaFija")),
            "tirea_fija": _f(r.get("tireaFija")), "vf_fija": _f(r.get("valorFinalFija")),
            "tamar_aplic": _f(r.get("tamarTnaAplicable")), "pata": r.get("pataElegida") or "",
            "spread_bps": _f(r.get("spreadTamarBps")), "volumen": _f(r.get("volumen"))}

def _es_valido_dual(r, hoy):
    return _filtro_comun(_f(r.get("precioSucio")), _f(r.get("volumen")),
                         _f(r.get("tireaTamar") or r.get("tireaFija")),
                         _d(r.get("fechaVencimiento")), hoy)

def build_duales(wb, hist, fecha_hoy, hora_str, fecha_dt, hora_dt):
    raw = api_get("/api/duales"); hoy = _d(fecha_hoy); liq = hoy + timedelta(days=1) if hoy else None
    rows = [_parse_dual(r) for r in raw if _es_valido_dual(r, hoy)]
    rows.sort(key=lambda r: r.get("vto") or date.max)
    print(f"  {len(rows)} válidos  ({len(raw)-len(rows)} excluidos)")
    if not rows: return
    snap = [{"ticker":r["ticker"],"dias_vto":r["dias_vto"],"precio":r["precio"],"var":r["var"],"volumen":r["volumen"],"tirea_tamar":r["tirea_tamar"],"tirea_fija":r["tirea_fija"]} for r in rows]
    add_snapshot_to_hist(hist, "DUAL", fecha_hoy, hora_str, snap)
    ws = wb.create_sheet("DUALES"); ws.sheet_properties.tabColor = TAB_DUAL
    write_meta(ws, "Duales — TAMAR vs Tasa Fija", fecha_dt, hora_dt, liq)
    write_hdr(ws, 5, DUAL_COLS)
    for i, r in enumerate(rows): write_row(ws, 6+i, DUAL_COLS, r, alt=(i%2==1))
    ws.freeze_panes = "A6"
    # ── Solapa de sensibilidad TAMAR + DUALES combinados ──
    print("  Construyendo sensibilidad DUAL/TAMAR...")
    raw_lecaps = api_get("/api/lecaps")
    raw_tamar  = api_get("/api/tamar")
    raw_duales = api_get("/api/duales")
    sens_rows, rates, ref_tirea = _build_sens_tamar_raw(raw_tamar, raw_duales, raw_lecaps, liq)
    if sens_rows:
        ws_s = wb.create_sheet("DUAL-TAMAR Sens"); ws_s.sheet_view.showGridLines = False
        _write_sens_sheet(ws_s, sens_rows, rates, ref_tirea,
                          "Sensibilidad TAMAR/DUALES — TIREA resultante por escenario",
                          TAB_DUAL, fecha_dt, hora_dt)

# ═══════════════════════════════════════════════════════════════════════════════
# 6-7. ONs
# ═══════════════════════════════════════════════════════════════════════════════
ONS_COLS = [
    {"key":"ticker",      "label":"Ticker",        "w": 9},
    {"key":"emisor",      "label":"Empresa",        "w":22},
    {"key":"vto",         "label":"Vencimiento",    "w":13, "fmt":"DD/MM/YYYY"},
    {"key":"precio_ars",  "label":"Precio ARS",     "w":13, "align":"right", "fmt":"#,##0"},
    {"key":"var",         "label":"% Var Diaria",   "w":13, "align":"right", "fmt":"0.00%"},
    {"key":"precio_d",    "label":"Precio D",       "w":13, "align":"right", "fmt":"#,##0.00"},
    {"key":"var_d",       "label":"% Var D",        "w":13, "align":"right", "fmt":"0.00%"},
    {"key":"tirea_d",     "label":"TIREA D (%)",    "w":14, "align":"right", "fmt":"0.00%"},
    {"key":"tna_d",       "label":"TNA D (%)",      "w":13, "align":"right", "fmt":"0.00%"},
    {"key":"md_d",        "label":"MD D",           "w":10, "align":"right", "fmt":"#,##0.00"},
    {"key":"precio_c",    "label":"Precio C",       "w":13, "align":"right", "fmt":"#,##0.00"},
    {"key":"var_c",       "label":"% Var C",        "w":13, "align":"right", "fmt":"0.00%"},
    {"key":"tirea_c",     "label":"TIREA C (%)",    "w":14, "align":"right", "fmt":"0.00%"},
    {"key":"tna_c",       "label":"TNA C (%)",      "w":13, "align":"right", "fmt":"0.00%"},
    {"key":"md_c",        "label":"MD C",           "w":10, "align":"right", "fmt":"#,##0.00"},
    {"key":"vt",          "label":"VT",             "w": 9, "align":"right", "fmt":"#,##0.00"},
    {"key":"paridad_d",   "label":"Paridad D",      "w":13, "align":"right", "fmt":"0.00%"},
    {"key":"ai",          "label":"Int. Corrido",   "w":13, "align":"right", "fmt":"#,##0.00"},
    {"key":"calificacion","label":"Calificación",   "w":14},
    {"key":"sector",      "label":"Sector",         "w":16},
    {"key":"volumen",     "label":"Volumen",        "w":18, "align":"right", "fmt":"#,##0"},
]

def _parse_ons(r, liq):
    p = _f(r.get("paridad") or r.get("paridadMep"))
    return {"ticker": (r.get("ticker") or "").strip(),
            "emisor": r.get("emisor") or r.get("empresa") or "",
            "vto": _d(r.get("fechaVencimiento") or r.get("vencimiento")),
            "precio_ars": _f(r.get("precioSucioArs") or r.get("precioArs")),
            "var":   _pct(r.get("variacionDiaria") or r.get("variacion_diaria")),
            "precio_d": _f(r.get("precioD") or r.get("precioMep")),
            "var_d":    _pct(r.get("variacionDiariaD") or r.get("variacionDiaria")),
            "tirea_d":  _f(r.get("tireaMep")),
            "tna_d":    _f(r.get("tnaMep")),
            "md_d":     _f(r.get("mdMep")),
            "precio_c": _f(r.get("precioC") or r.get("precioCcl")),
            "var_c":    _pct(r.get("variacionDiariaC") or r.get("variacionDiaria")),
            "tirea_c":  _f(r.get("tireaCcl")),
            "tna_c":    _f(r.get("tnaCcl")),
            "md_c":     _f(r.get("mdCcl")),
            "vt":       _f(r.get("vt") or r.get("valorTecnico")),
            "paridad_d": p if p and p <= 3 else (p/100 if p else None),
            "ai":       _f(r.get("ai") or r.get("interesesCorridos")),
            "calificacion": r.get("calificacionRiesgo") or r.get("calificacion") or "",
            "sector":   r.get("sector") or "",
            "volumen":  _f(r.get("volumen") or r.get("v"))}

def _es_valido_ons(r, hoy):
    return _filtro_comun(_f(r.get("precioD") or r.get("precioMep") or r.get("precioMEP")),
                         _f(r.get("volumen") or r.get("v")),
                         _f(r.get("tireaMep") or r.get("tireaMEP")),
                         _d(r.get("fechaVencimiento") or r.get("vencimiento")), hoy)

def build_ons(wb, hist, endpoint, instr_key, sheet_name, titulo, tab_color,
              fecha_hoy, hora_str, fecha_dt, hora_dt):
    raw = api_get(endpoint); hoy = _d(fecha_hoy); liq = hoy + timedelta(days=1) if hoy else None
    rows = [_parse_ons(r, liq) for r in raw if _es_valido_ons(r, hoy)]
    rows.sort(key=lambda r: r.get("vto") or date.max)
    print(f"  {len(rows)} válidos  ({len(raw)-len(rows)} excluidos)")
    if not rows: return
    snap = [{"ticker":r["ticker"],"md":r["md_d"],"precio":r["precio_d"],"var":r["var"],"volumen":r["volumen"],"tirea":r["tirea_d"],"tirea_ccl":r["tirea_c"]} for r in rows]
    add_snapshot_to_hist(hist, instr_key, fecha_hoy, hora_str, snap)
    ws = wb.create_sheet(sheet_name); ws.sheet_properties.tabColor = tab_color
    write_meta(ws, titulo, fecha_dt, hora_dt, liq)
    write_hdr(ws, 5, ONS_COLS)
    for i, r in enumerate(rows): write_row(ws, 6+i, ONS_COLS, r, alt=(i%2==1))
    ws.freeze_panes = "A6"

# ═══════════════════════════════════════════════════════════════════════════════
# 8. HD SOBERANOS
# ═══════════════════════════════════════════════════════════════════════════════
HDSOB_COLS = [
    {"key":"ticker",    "label":"Ticker",      "w": 9},
    {"key":"categoria", "label":"Categoría",   "w":12},
    {"key":"vto",       "label":"Vencimiento", "w":13, "fmt":"DD/MM/YYYY"},
    {"key":"precio",  "label":"Precio",       "w":10, "align":"right", "fmt":"#,##0.00"},
    {"key":"var",     "label":"% Var Diaria", "w":11, "align":"right", "fmt":"0.00%"},
    {"key":"tirea",   "label":"TIREA (%)",    "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"tna",     "label":"TNA (%)",      "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"md",      "label":"MD",           "w": 8, "align":"right", "fmt":"0.00"},
    {"key":"vt",      "label":"VT",           "w":10, "align":"right", "fmt":"#,##0.00"},
    {"key":"paridad", "label":"Paridad",      "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"ai",      "label":"Int. Corrido", "w":12, "align":"right", "fmt":"#,##0.000"},
    {"key":"volumen", "label":"Volumen",      "w":15, "align":"right", "fmt":"#,##0"},
]

def _categoria_hdsob(r):
    emisor = (r.get("emisor") or "").strip().lower()
    ley    = (r.get("ley")    or "").strip().lower()
    if "bcra" in emisor: return "Bopreales"
    if "gobierno nacional" in emisor or "gobierno" in emisor:
        if "nueva york" in ley or "new york" in ley or "york" in ley: return "Globales"
        if "nacional" in ley or "argentina" in ley: return "Bonares"
        t = (r.get("ticker") or "").upper()
        if t.startswith("GD") or t.startswith("GE") or t.startswith("AE"): return "Globales"
        if t.startswith("AL") or t.startswith("AO") or t.startswith("AN") or t.startswith("AC"): return "Bonares"
    t = (r.get("ticker") or "").upper()
    if t.startswith("BP"): return "Bopreales"
    if t.startswith("GD") or t.startswith("GE") or t.startswith("AE"): return "Globales"
    if t.startswith("AL") or t.startswith("AO") or t.startswith("AN") or t.startswith("AC"): return "Bonares"
    return "Otros"

_CAT_ORDER = {"Globales": 0, "Bonares": 1, "Bopreales": 2, "Otros": 3}

def _parse_hdsob(r):
    return {"ticker": (r.get("ticker") or "").strip(),
            "categoria": _categoria_hdsob(r),
            "vto": _d(r.get("fechaVencimiento")), "precio": _f(r.get("precioSucio")),
            "var": _pct(r.get("variacionDiaria")), "tirea": _f(r.get("tirea")),
            "tna": _f(r.get("tna")), "md": _f(r.get("duracionMod")),
            "vt": _f(r.get("valorTecnico")), "paridad": _f(r.get("paridad")),
            "ai": _f(r.get("ai")), "volumen": _f(r.get("volumen"))}

def _es_valido_hdsob(r, hoy):
    return _filtro_comun(_f(r.get("precioSucio")), _f(r.get("volumen")), _f(r.get("tirea")),
                         _d(r.get("fechaVencimiento")), hoy)

def build_hdsob(wb, hist, fecha_hoy, hora_str, fecha_dt, hora_dt):
    raw = api_get("/api/bonoshd"); hoy = _d(fecha_hoy); liq = hoy + timedelta(days=1) if hoy else None
    rows = [_parse_hdsob(r) for r in raw if _es_valido_hdsob(r, hoy)]
    rows.sort(key=lambda r: (_CAT_ORDER.get(r.get("categoria","Otros"), 3), r.get("vto") or date.max))
    print(f"  {len(rows)} válidos  ({len(raw)-len(rows)} excluidos)")
    if not rows: return
    snap = [{"ticker":r["ticker"],"md":r["md"],"precio":r["precio"],"var":r["var"],"volumen":r["volumen"],"tirea":r["tirea"],"tna":r["tna"]} for r in rows]
    add_snapshot_to_hist(hist, "HDSOB", fecha_hoy, hora_str, snap)
    ws = wb.create_sheet("HD_Sob"); ws.sheet_properties.tabColor = TAB_HDSOB
    write_meta(ws, "Bonos Hard Dollar Soberanos (Bonares + Globales)", fecha_dt, hora_dt, liq)
    write_hdr(ws, 5, HDSOB_COLS)
    for i, r in enumerate(rows): write_row(ws, 6+i, HDSOB_COLS, r, alt=(i%2==1))
    ws.freeze_panes = "A6"

# ═══════════════════════════════════════════════════════════════════════════════
# 9. BOTE
# ═══════════════════════════════════════════════════════════════════════════════
BOTE_COLS = [
    {"key":"ticker",        "label":"Ticker",        "w": 9},
    {"key":"vto",           "label":"Vencimiento",   "w":13, "fmt":"DD/MM/YYYY"},
    {"key":"precio_sucio",  "label":"Precio Sucio",  "w":12, "align":"right", "fmt":"#,##0.00"},
    {"key":"precio_limpio", "label":"Precio Limpio", "w":12, "align":"right", "fmt":"#,##0.00"},
    {"key":"var",           "label":"% Var Diaria",  "w":11, "align":"right", "fmt":"0.00%"},
    {"key":"tirea",         "label":"TIREA (%)",     "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"tna",           "label":"TNA (%)",       "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"md",            "label":"MD",            "w": 8, "align":"right", "fmt":"0.00"},
    {"key":"wal",           "label":"WAL",           "w": 8, "align":"right", "fmt":"0.00"},
    {"key":"vt",            "label":"VT",            "w":10, "align":"right", "fmt":"#,##0.00"},
    {"key":"vr",            "label":"VR",            "w": 8, "align":"right", "fmt":"#,##0.00"},
    {"key":"paridad",       "label":"Paridad",       "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"ai",            "label":"Int. Corrido",  "w":12, "align":"right", "fmt":"#,##0.000"},
    {"key":"cupon",         "label":"Cupón %",       "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"volumen",       "label":"Volumen",       "w":15, "align":"right", "fmt":"#,##0"},
]

def _parse_bote(r):
    return {"ticker": (r.get("ticker") or "").strip(), "emisor": r.get("emisor") or "",
            "vto": _d(r.get("fechaVencimiento")),
            "precio_sucio": _f(r.get("precioSucio")), "precio_limpio": _f(r.get("precioLimpio")),
            "var": _pct(r.get("variacionDiaria")), "tirea": _f(r.get("tirea")),
            "tna": _f(r.get("tna")), "md": _f(r.get("duracionMod")), "wal": _f(r.get("wal")),
            "vt": _f(r.get("valorTecnico")), "vr": _f(r.get("vr")),
            "paridad": _f(r.get("paridad")), "ai": _f(r.get("ai")),
            "cupon": _f(r.get("cuponFijoPct")), "volumen": _f(r.get("volumen"))}

def _es_valido_bote(r, hoy):
    return _filtro_comun(_f(r.get("precioSucio")), _f(r.get("volumen")), _f(r.get("tirea")),
                         _d(r.get("fechaVencimiento")), hoy)

def build_bote(wb, hist, fecha_hoy, hora_str, fecha_dt, hora_dt):
    raw = api_get("/api/bote"); hoy = _d(fecha_hoy); liq = hoy + timedelta(days=1) if hoy else None
    rows = [_parse_bote(r) for r in raw if _es_valido_bote(r, hoy)]
    rows.sort(key=lambda r: r.get("vto") or date.max)
    print(f"  {len(rows)} válidos  ({len(raw)-len(rows)} excluidos)")
    if not rows: return
    snap = [{"ticker":r["ticker"],"md":r["md"],"precio_sucio":r["precio_sucio"],"var":r["var"],"volumen":r["volumen"],"tirea":r["tirea"],"tna":r["tna"],"wal":r["wal"]} for r in rows]
    add_snapshot_to_hist(hist, "BOTE", fecha_hoy, hora_str, snap)
    ws = wb.create_sheet("BOTE"); ws.sheet_properties.tabColor = TAB_BOTE
    write_meta(ws, "BOTE — Bonos Tasa Fija USD", fecha_dt, hora_dt, liq)
    write_hdr(ws, 5, BOTE_COLS)
    for i, r in enumerate(rows): write_row(ws, 6+i, BOTE_COLS, r, alt=(i%2==1))
    ws.freeze_panes = "A6"

# ═══════════════════════════════════════════════════════════════════════════════
# HISTORIA  v3
# ═══════════════════════════════════════════════════════════════════════════════
def build_historia(wb, hist):
    ws = wb.create_sheet("HISTORIA")
    ws.sheet_properties.tabColor = TAB_HIST
    ws.sheet_view.showGridLines   = False
    for r in range(1, 3000):
        for c in range(1, 14):
            ws.cell(r, c).fill = _fill(BG_TITLE)

    ws["A1"].value = "Historia de cierres diarios  v3.0"
    ws["A1"].font  = Font(name="Arial", bold=True, color=FG_TITLE, size=12)
    ws["A1"].fill  = _fill(BG_TITLE)
    ws["A2"].value = "Extra: DUAL=TIREA Fija  |  ONs=TIREA CCL  |  BOTE=WAL"
    ws["A2"].font  = _font(FG_MUTED, size=7); ws["A2"].fill = _fill(BG_TITLE)

    hdrs   = ["Instrumento","Fecha","Hora","Ticker"] + HIST_FIXED_HDRS
    widths = [14, 11, 7, 10] + HIST_FIXED_WIDTHS
    for i, (h, w) in enumerate(zip(hdrs, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(3, i)
        c.value = h; c.fill = _fill(BG_HDR); c.font = _font(FG_WHITE, bold=True)
        c.alignment = _center(); c.border = _border()

    all_entries = []
    for instr, entries in hist.items():
        m = HIST_MAP.get(instr)
        if not m: continue
        for entry in entries:
            all_entries.append((instr, entry))
    all_entries.sort(key=lambda x: (x[1].get("fecha",""), x[0]), reverse=True)

    row_idx = 4
    for instr, entry in all_entries:
        m = HIST_MAP.get(instr)
        if not m: continue
        fecha     = entry.get("fecha","")
        hora      = entry.get("hora","")
        snap_rows = entry.get("snap") or []
        raw_rows  = entry.get("_raw") or []
        use_old   = not snap_rows and bool(raw_rows)
        rows_src  = snap_rows if not use_old else raw_rows
        for r in rows_src:
            if use_old:
                ticker  = r.get("_ticker","")
                hora_r  = str(r.get("_hora") or hora or "")
                vals_raw = r.get("_vals") or []
                v_dim   = vals_raw[0] if len(vals_raw) > 0 else None
                v_precio = None
                v_var    = None
                v_vol    = None
                v_tem   = vals_raw[1] if len(vals_raw) > 1 else None
                v_tna   = vals_raw[2] if len(vals_raw) > 2 else None
                v_tirea = vals_raw[3] if len(vals_raw) > 3 else None
                v_extra = vals_raw[4] if len(vals_raw) > 4 else None
            else:
                ticker   = r.get("ticker","")
                hora_r   = hora
                v_dim    = r.get(m["dim"])
                v_precio = r.get(m["precio"]) if m.get("precio") else None
                v_var    = r.get(m["var"])    if m.get("var")    else None
                v_vol    = r.get(m["vol"])    if m.get("vol")    else None
                v_tem    = r.get(m["tem"])    if m["tem"]        else None
                v_tna    = r.get(m["tna"])    if m["tna"]        else None
                v_tirea  = r.get(m["tirea"])  if m["tirea"]      else None
                v_extra  = r.get(m["extra"])  if m["extra"]      else None

            ws.cell(row_idx,1).value = instr
            ws.cell(row_idx,2).value = fecha
            ws.cell(row_idx,3).value = hora_r
            ws.cell(row_idx,4).value = ticker

            c5 = ws.cell(row_idx,5); c5.value = v_precio
            if v_precio is not None: c5.number_format = m["precio_fmt"]
            c6 = ws.cell(row_idx,6); c6.value = v_var
            if v_var is not None: c6.number_format = "0.00%"
            c7 = ws.cell(row_idx,7); c7.value = v_tem
            if v_tem is not None: c7.number_format = "0.00%"
            c8 = ws.cell(row_idx,8); c8.value = v_tna
            if v_tna is not None: c8.number_format = "0.00%"
            c9 = ws.cell(row_idx,9); c9.value = v_tirea
            if v_tirea is not None: c9.number_format = "0.00%"
            c10 = ws.cell(row_idx,10); c10.value = v_dim
            if v_dim is not None: c10.number_format = m["dim_fmt"]
            c11 = ws.cell(row_idx,11); c11.value = v_vol
            if v_vol is not None: c11.number_format = "#,##0"
            c12 = ws.cell(row_idx,12); c12.value = v_extra
            if v_extra is not None and m["extra_fmt"]:
                c12.number_format = m["extra_fmt"]

            alt = (row_idx % 2 == 0)
            for ci in range(1, 13):
                c = ws.cell(row_idx, ci)
                c.fill = _fill(BG_ALT if alt else BG_ROW)
                c.font = _font(FG_DATA, size=8)
                c.border = _border(); c.alignment = _center()
            row_idx += 1

    ws.freeze_panes = "A4"
    print(f"  {row_idx-4} registros en historia")


# ═══════════════════════════════════════════════════════════════════════════════
# DÓLAR
# ═══════════════════════════════════════════════════════════════════════════════
DOLAR_URL = "https://api.alquimiaconsultora.com/output/dolares.json"
BCRA_URL  = "https://api.alquimiaconsultora.com/output/bcra.json"

DOLAR_MAIN_COLS = [
    {"key": "tipo",      "label": "Tipo",       "w": 16},
    {"key": "precio",    "label": "Precio",     "w": 12, "align": "right", "fmt": "#,##0.00"},
    {"key": "compra",    "label": "Compra",     "w": 12, "align": "right", "fmt": "#,##0.00"},
    {"key": "venta",     "label": "Venta",      "w": 12, "align": "right", "fmt": "#,##0.00"},
    {"key": "variacion", "label": "Var %",      "w":  9, "align": "right", "fmt": "0.00%"},
]
DOLAR_MEP_CCL_COLS = [
    {"key": "tipo",      "label": "Tipo",       "w": 16},
    {"key": "bono",      "label": "Bono",       "w": 10},
    {"key": "plazo",     "label": "Plazo",      "w":  8},
    {"key": "precio",    "label": "Precio",     "w": 12, "align": "right", "fmt": "#,##0.00"},
    {"key": "variacion", "label": "Var %",      "w":  9, "align": "right", "fmt": "0.00%"},
]
DOLAR_CRIPTO_COLS = [
    {"key": "tipo",      "label": "Cripto",     "w": 10},
    {"key": "compra",    "label": "Compra",     "w": 12, "align": "right", "fmt": "#,##0.00"},
    {"key": "venta",     "label": "Venta",      "w": 12, "align": "right", "fmt": "#,##0.00"},
    {"key": "variacion", "label": "Var %",      "w":  9, "align": "right", "fmt": "0.00%"},
]

def _pct_direct(v):
    x = _f(v)
    return None if x is None else x / 100.0

def _get_mayorista_a3500():
    """Obtiene el dólar mayorista A3500 desde bcra.json.
    Estructura real: { "series": { "A3500": { "YYYY-MM-DD": valor, ... }, ... } }
    Devuelve (precio, variacion_pct) usando los dos últimos valores disponibles.
    """
    try:
        r = requests.get(BCRA_URL, timeout=30); r.raise_for_status()
        data = r.json()
        serie = data.get("series", {}).get("A3500")
        if not serie or not isinstance(serie, dict):
            print("  [WARN] Clave 'A3500' no encontrada o vacía en bcra.json")
            return None, None
        fechas_ordenadas = sorted(serie.keys())
        ultima_fecha     = fechas_ordenadas[-1]
        valor            = _f(serie[ultima_fecha])
        # Variación: (ultimo - anteultimo) / anteultimo
        variacion = None
        if len(fechas_ordenadas) >= 2:
            anteultima_fecha = fechas_ordenadas[-2]
            valor_ant        = _f(serie[anteultima_fecha])
            if valor_ant and valor_ant != 0 and valor is not None:
                variacion = (valor - valor_ant) / valor_ant
        print(f"  A3500 fecha más reciente: {ultima_fecha} → {valor}  |  var: {variacion}")
        return valor, variacion
    except Exception as e:
        print(f"  [WARN] No se pudo obtener A3500 de bcra.json: {e}")
    return None, None

def build_dolar(wb, fecha_dt, hora_dt):
    print(f"  Consultando {DOLAR_URL}...")
    try:
        r = requests.get(DOLAR_URL, timeout=30); r.raise_for_status()
        payload = r.json()
    except Exception as e:
        print(f"  [ERROR] No se pudo obtener datos de dólar: {e}"); return

    data = payload.get("data", {})
    if not data:
        print("  [WARN] Respuesta vacía, se omite solapa DÓLAR."); return

    print(f"  Consultando {BCRA_URL} (A3500 mayorista)...")
    mayorista_precio, mayorista_var = _get_mayorista_a3500()
    if mayorista_precio is not None:
        print(f"  A3500 mayorista: {mayorista_precio}")
    else:
        print("  [WARN] No se encontró el A3500 en bcra.json")

    try:
        ws = wb.create_sheet("DÓLAR")
    except Exception:
        ws = wb.create_sheet("DOLAR")
    ws.sheet_properties.tabColor = TAB_DOLAR

    try:
        write_meta(ws, "Dólar — Cotizaciones del Día", fecha_dt, hora_dt, None)

        # Tipos simples: Mayorista ya no viene de dolares.json
        TIPOS_SIMPLES = [
            ("Oficial", "oficial"),
            ("Ahorro", "ahorro"), ("Tarjeta", "tarjeta"), ("Blue", "blue"),
        ]
        r5 = ws.cell(5, 1); r5.value = "Cotizaciones Generales"
        r5.font = _font(FG_TITLE, bold=True, size=9); r5.fill = _fill(BG_TITLE)
        ws.row_dimensions[5].height = 14

        write_hdr(ws, 6, DOLAR_MAIN_COLS)
        row_idx = 7

        # Mayorista desde bcra.json (A3500) — va primero
        fila_mayorista = {
            "tipo":      "Mayorista (A3500)",
            "precio":    mayorista_precio,
            "compra":    None,
            "venta":     None,
            "variacion": mayorista_var,
        }
        write_row(ws, row_idx, DOLAR_MAIN_COLS, fila_mayorista, alt=(row_idx % 2 == 0))
        row_idx += 1

        for nombre, clave in TIPOS_SIMPLES:
            d = data.get(clave)
            if not d: continue
            fila = {"tipo": nombre, "precio": _f(d.get("price")),
                    "compra": _f(d.get("bid")), "venta": _f(d.get("ask")),
                    "variacion": _pct_direct(d.get("variation"))}
            write_row(ws, row_idx, DOLAR_MAIN_COLS, fila, alt=(row_idx % 2 == 0))
            row_idx += 1

        row_idx += 1
        r_sub = ws.cell(row_idx, 1); r_sub.value = "Dólar Cripto"
        r_sub.font = _font(FG_TITLE, bold=True, size=9); r_sub.fill = _fill(BG_TITLE)
        ws.row_dimensions[row_idx].height = 14; row_idx += 1

        write_hdr(ws, row_idx, DOLAR_CRIPTO_COLS); row_idx += 1
        cripto = data.get("cripto", {})
        for clave in ["ccb", "usdt", "usdc"]:
            d = cripto.get(clave)
            if not d: continue
            fila = {"tipo": clave.upper(), "compra": _f(d.get("bid")),
                    "venta": _f(d.get("ask")), "variacion": _pct_direct(d.get("variation"))}
            write_row(ws, row_idx, DOLAR_CRIPTO_COLS, fila, alt=(row_idx % 2 == 0))
            row_idx += 1

        row_idx += 1
        r_sub2 = ws.cell(row_idx, 1); r_sub2.value = "MEP / CCL"
        r_sub2.font = _font(FG_TITLE, bold=True, size=9); r_sub2.fill = _fill(BG_TITLE)
        ws.row_dimensions[row_idx].height = 14; row_idx += 1

        write_hdr(ws, row_idx, DOLAR_MEP_CCL_COLS); row_idx += 1
        for tipo_key, tipo_lbl in [("mep", "MEP"), ("ccl", "CCL")]:
            grupo = data.get(tipo_key, {})
            for bono_key in ["al30", "gd30", "letras", "bpo27"]:
                d_bono = grupo.get(bono_key)
                if not d_bono: continue
                for plazo_key in ["24hs", "ci"]:
                    d_plazo = d_bono.get(plazo_key)
                    if not d_plazo: continue
                    fila = {"tipo": tipo_lbl, "bono": bono_key.upper(), "plazo": plazo_key,
                            "precio": _f(d_plazo.get("price")),
                            "variacion": _pct_direct(d_plazo.get("variation"))}
                    write_row(ws, row_idx, DOLAR_MEP_CCL_COLS, fila, alt=(row_idx % 2 == 0))
                    row_idx += 1

        ws.freeze_panes = "A7"
        print(f"  Solapa DÓLAR creada ({row_idx - 7} filas de datos)")
    except Exception as e:
        import traceback
        print(f"  [ERROR] Falló al construir solapa DÓLAR: {e}")
        traceback.print_exc()


# ═══════════════════════════════════════════════════════════════════════════════
# ACCIONES ARGENTINAS
# ═══════════════════════════════════════════════════════════════════════════════
ACCIONES_COLS = [
    {"key": "ticker",          "label": "Ticker",        "w":  9},
    {"key": "empresa",         "label": "Empresa",       "w": 24},
    {"key": "sector",          "label": "Sector",        "w": 16},
    {"key": "precio",          "label": "Precio",        "w": 12, "align": "right", "fmt": "#,##0.00"},
    {"key": "variacionDiaria", "label": "Var % Día",     "w": 10, "align": "right", "fmt": "0.00%"},
    {"key": "varYtd",          "label": "YTD %",         "w": 10, "align": "right", "fmt": "0.00%"},
    {"key": "volumen",         "label": "Volumen",       "w": 16, "align": "right", "fmt": "#,##0"},
    {"key": "marketCapUsd",    "label": "Market Cap USD","w": 18, "align": "right", "fmt": "#,##0"},
    {"key": "pe",              "label": "P/E",           "w":  9, "align": "right", "fmt": "#,##0.00"},
    {"key": "roa",             "label": "ROA %",         "w":  9, "align": "right", "fmt": "0.00%"},
    {"key": "roe",             "label": "ROE %",         "w":  9, "align": "right", "fmt": "0.00%"},
    {"key": "pn",              "label": "Patr. Neto",    "w": 14, "align": "right", "fmt": "#,##0"},
]

def build_acciones(wb, fecha_dt, hora_dt):
    url = f"{API_BASE}/api/acciones"
    print(f"  Consultando {url}...")
    try:
        r = requests.get(url, timeout=30); r.raise_for_status()
        raw = r.json()
        if isinstance(raw, dict):
            raw = raw.get("rows", raw.get("data", []))
    except Exception as e:
        print(f"  [ERROR] No se pudo obtener acciones argentinas: {e}"); return

    if not raw:
        print("  [WARN] Sin datos de acciones argentinas."); return

    rows = []
    for d in raw:
        rows.append({
            "ticker":          d.get("ticker", ""),
            "empresa":         d.get("empresa", ""),
            "sector":          d.get("sector", ""),
            "precio":          _f(d.get("precio")),
            "variacionDiaria": _pct_direct(d.get("variacionDiaria")),
            "varYtd":          _f(d.get("varYtd")),
            "volumen":         _f(d.get("volumen")),
            "marketCapUsd":    _f(d.get("marketCapUsd")),
            "pe":              _f(d.get("pe")),
            "roa":             _pct_direct(d.get("roa")),
            "roe":             _pct_direct(d.get("roe")),
            "pn":              _f(d.get("pn")),
        })

    rows.sort(key=lambda r: r.get("ticker") or "")

    ws = wb.create_sheet("ACCIONES AR")
    ws.sheet_properties.tabColor = TAB_ACCIONES
    write_meta(ws, "Acciones Argentinas — Mercado Local", fecha_dt, hora_dt, None)
    write_hdr(ws, 5, ACCIONES_COLS)
    for i, row in enumerate(rows):
        write_row(ws, 6 + i, ACCIONES_COLS, row, alt=(i % 2 == 1))
    ws.freeze_panes = "A6"
    print(f"  Solapa ACCIONES AR creada ({len(rows)} registros)")


# ═══════════════════════════════════════════════════════════════════════════════
# ADRs ARGENTINA
# ═══════════════════════════════════════════════════════════════════════════════
ADR_URL = "https://api.alquimiaconsultora.com/output/yahoo_adrs_arg.json"

ADR_COLS = [
    {"key": "symbol",       "label": "Ticker",      "w":  8},
    {"key": "name",         "label": "Nombre",       "w": 28},
    {"key": "sector",       "label": "Sector",       "w": 20},
    {"key": "price",        "label": "Precio USD",  "w": 12, "align": "right", "fmt": "#,##0.00"},
    {"key": "change_pct",   "label": "Var %",        "w":  9, "align": "right", "fmt": "0.00%"},
    {"key": "day_high",     "label": "Máx. Día",    "w": 11, "align": "right", "fmt": "#,##0.00"},
    {"key": "day_low",      "label": "Mín. Día",    "w": 11, "align": "right", "fmt": "#,##0.00"},
    {"key": "volume",       "label": "Volumen",      "w": 14, "align": "right", "fmt": "#,##0"},
    {"key": "market_cap",   "label": "Market Cap",   "w": 16, "align": "right", "fmt": "#,##0"},
    {"key": "pe",           "label": "P/E",          "w":  8, "align": "right", "fmt": "#,##0.00"},
    {"key": "roa_pct",      "label": "ROA %",        "w":  9, "align": "right", "fmt": "0.00%"},
    {"key": "roe_pct",      "label": "ROE %",        "w":  9, "align": "right", "fmt": "0.00%"},
    {"key": "ytd_change",   "label": "YTD %",        "w":  9, "align": "right", "fmt": "0.00%"},
]

def build_adr(wb, fecha_dt, hora_dt):
    print(f"  Consultando {ADR_URL}...")
    try:
        r = requests.get(ADR_URL, timeout=30); r.raise_for_status()
        payload = r.json()
    except Exception as e:
        print(f"  [ERROR] No se pudo obtener datos de ADRs: {e}"); return

    raw = payload.get("adrs_argentina", [])
    if not raw:
        print("  [WARN] Sin datos de ADRs."); return

    rows = []
    for d in raw:
        rows.append({
            "symbol":     d.get("symbol", ""),
            "name":       d.get("name", ""),
            "sector":     d.get("sector", ""),
            "price":      _f(d.get("price")),
            "change_pct": _pct_direct(d.get("change_pct")),
            "day_high":   _f(d.get("day_high")),
            "day_low":    _f(d.get("day_low")),
            "volume":     _f(d.get("volume")),
            "market_cap": _f(d.get("market_cap_usd")),
            "pe":         _f(d.get("pe")),
            "roa_pct":    _pct_direct(d.get("roa_pct")),
            "roe_pct":    _pct_direct(d.get("roe_pct")),
            "ytd_change": _pct_direct(d.get("ytd_change_pct")),
        })

    rows.sort(key=lambda r: r.get("symbol") or "")

    try:
        ws = wb.create_sheet("ADR")
    except Exception:
        ws = wb.create_sheet("ADR")
    ws.sheet_properties.tabColor = TAB_ADR

    write_meta(ws, "ADRs Argentina — Cotización NYSE", fecha_dt, hora_dt, None)
    write_hdr(ws, 5, ADR_COLS)
    for i, row in enumerate(rows):
        write_row(ws, 6 + i, ADR_COLS, row, alt=(i % 2 == 1))
    ws.freeze_panes = "A6"
    print(f"  Solapa ADR creada ({len(rows)} registros)")


# ═══════════════════════════════════════════════════════════════════════════════
# NYSE / US Large Caps
# ═══════════════════════════════════════════════════════════════════════════════
NYSE_URL = "https://api.alquimiaconsultora.com/output/yahoo_nyse.json"

NYSE_COLS = [
    {"key": "symbol",       "label": "Ticker",      "w":  8},
    {"key": "name",         "label": "Nombre",       "w": 38},
    {"key": "sector",       "label": "Sector",       "w": 22},
    {"key": "price",        "label": "Precio USD",  "w": 12, "align": "right", "fmt": "#,##0.00"},
    {"key": "change_pct",   "label": "Var %",        "w":  9, "align": "right", "fmt": "0.00%"},
    {"key": "day_high",     "label": "Máx. Día",    "w": 11, "align": "right", "fmt": "#,##0.00"},
    {"key": "day_low",      "label": "Mín. Día",    "w": 11, "align": "right", "fmt": "#,##0.00"},
    {"key": "volume",       "label": "Volumen",      "w": 16, "align": "right", "fmt": "#,##0"},
    {"key": "market_cap",   "label": "Market Cap",   "w": 18, "align": "right", "fmt": "#,##0"},
    {"key": "pe",           "label": "P/E",          "w":  8, "align": "right", "fmt": "#,##0.00"},
    {"key": "roa_pct",      "label": "ROA %",        "w":  9, "align": "right", "fmt": "0.00%"},
    {"key": "roe_pct",      "label": "ROE %",        "w":  9, "align": "right", "fmt": "0.00%"},
    {"key": "ytd_change",   "label": "YTD %",        "w":  9, "align": "right", "fmt": "0.00%"},
]

def build_nyse(wb, fecha_dt, hora_dt):
    print(f"  Consultando {NYSE_URL}...")
    try:
        r = requests.get(NYSE_URL, timeout=30); r.raise_for_status()
        payload = r.json()
    except Exception as e:
        print(f"  [ERROR] No se pudo obtener datos de NYSE: {e}"); return

    raw = payload.get("nyse", [])
    if not raw:
        print("  [WARN] Sin datos de NYSE."); return

    rows = []
    for d in raw:
        rows.append({
            "symbol":     d.get("symbol", ""),
            "name":       d.get("long_name") or d.get("name", ""),
            "sector":     d.get("sector", ""),
            "price":      _f(d.get("price")),
            "change_pct": _pct_direct(d.get("change_pct")),
            "day_high":   _f(d.get("day_high")),
            "day_low":    _f(d.get("day_low")),
            "volume":     _f(d.get("volume")),
            "market_cap": _f(d.get("market_cap_usd")),
            "pe":         _f(d.get("pe")),
            "roa_pct":    _pct_direct(d.get("roa_pct")),
            "roe_pct":    _pct_direct(d.get("roe_pct")),
            "ytd_change": _pct_direct(d.get("ytd_change_pct")),
        })

    rows.sort(key=lambda r: r.get("symbol") or "")

    try:
        ws = wb.create_sheet("NYSE")
    except Exception:
        ws = wb.create_sheet("NYSE")
    ws.sheet_properties.tabColor = TAB_NYSE

    write_meta(ws, "NYSE / US Large Caps — Yahoo Finance", fecha_dt, hora_dt, None)
    write_hdr(ws, 5, NYSE_COLS)
    for i, row in enumerate(rows):
        write_row(ws, 6 + i, NYSE_COLS, row, alt=(i % 2 == 1))
    ws.freeze_panes = "A6"
    print(f"  Solapa NYSE creada ({len(rows)} registros)")


# ═══════════════════════════════════════════════════════════════════════════════
# CEDEARs / PUSH PRICES
# ═══════════════════════════════════════════════════════════════════════════════
CEDEAR_URL = "https://api.alquimiaconsultora.com/output/push_prices.json"

CEDEAR_COLS = [
    {"key": "ticker",     "label": "Ticker",      "w": 12},
    {"key": "price",      "label": "Precio",      "w": 12, "align": "right", "fmt": "#,##0.000"},
    {"key": "pct_change", "label": "Var %",       "w":  9, "align": "right", "fmt": "0.00%"},
    {"key": "q_op",       "label": "Cant. Op.",   "w": 10, "align": "right", "fmt": "#,##0"},
    {"key": "volumen",    "label": "Volumen",      "w": 18, "align": "right", "fmt": "#,##0.00"},
    {"key": "fuente",     "label": "Fuente",       "w":  8},
    {"key": "hora",       "label": "Hora",         "w":  8},
]

def build_cedear(wb, fecha_dt, hora_dt):
    print(f"  Consultando {CEDEAR_URL}...")
    try:
        r = requests.get(CEDEAR_URL, timeout=60); r.raise_for_status()
        raw = r.json()
    except Exception as e:
        print(f"  [ERROR] No se pudo obtener datos de CEDEARs: {e}"); return

    if not isinstance(raw, list):
        print("  [WARN] Formato inesperado en push_prices."); return

    rows = []
    for d in raw:
        v = _f(d.get("v"))
        if not v or v <= 0: continue
        rows.append({
            "ticker":     d.get("ticker", ""),
            "price":      _f(d.get("price")),
            "pct_change": _pct_direct(d.get("pct_change")),
            "q_op":       int(d.get("q_op") or 0),
            "volumen":    v,
            "fuente":     d.get("fuente", ""),
            "hora":       d.get("hora_input", ""),
        })

    rows.sort(key=lambda r: r.get("ticker") or "")

    try:
        ws = wb.create_sheet("CEDEAR")
    except Exception:
        ws = wb.create_sheet("CEDEAR")
    ws.sheet_properties.tabColor = TAB_CEDEAR

    write_meta(ws, "CEDEARs / Push Prices — Mercado Local", fecha_dt, hora_dt, None)
    write_hdr(ws, 5, CEDEAR_COLS)
    for i, row in enumerate(rows):
        write_row(ws, 6 + i, CEDEAR_COLS, row, alt=(i % 2 == 1))
    ws.freeze_panes = "A6"
    print(f"  Solapa CEDEAR creada ({len(rows)} registros con volumen)")


# ═══════════════════════════════════════════════════════════════════════════════
# MERCADOS INTERNACIONALES (yahoo_markets.json)
# ═══════════════════════════════════════════════════════════════════════════════
MARKETS_URL = "https://api.alquimiaconsultora.com/output/yahoo_markets.json"

MKT_COLS = [
    {"key": "symbol",     "label": "Symbol",     "w": 14},
    {"key": "name",       "label": "Nombre",     "w": 26},
    {"key": "price",      "label": "Precio",     "w": 14, "align": "right", "fmt": "#,##0.0000"},
    {"key": "change",     "label": "Cambio",     "w": 14, "align": "right", "fmt": "#,##0.0000"},
    {"key": "change_pct", "label": "Cambio %",   "w": 10, "align": "right", "fmt": "0.00%"},
]

_US_SYMBOLS    = {"^VIX", "DX-Y.NYB", "^DJI", "^GSPC", "^IXIC", "^RUT"}
_AM_SYMBOLS    = {"^BVSP", "^MERV", "^GSPTSE"}
_EU_SYMBOLS    = {"^FTSE", "^GDAXI", "^FCHI", "^STOXX50E", "^IBEX"}
_ASIA_SYMBOLS  = {"^N225", "^HSI", "^KS11", "^AXJO", "000001.SS"}


def _parse_mkt_row(d):
    return {
        "symbol":     d.get("symbol", ""),
        "name":       d.get("name", d.get("label", "")),
        "price":      _f(d.get("price")),
        "change":     _f(d.get("change")),
        "change_pct": _pct_direct(d.get("change_pct")),
    }


def _build_mkt_sheet(wb, tab_name, tab_color, title, rows, fecha_dt, hora_dt):
    try:
        ws = wb.create_sheet(tab_name)
    except Exception:
        ws = wb.create_sheet(tab_name + "_")
    ws.sheet_properties.tabColor = tab_color
    write_meta(ws, title, fecha_dt, hora_dt, None)
    write_hdr(ws, 5, MKT_COLS)
    for i, row in enumerate(rows):
        write_row(ws, 6 + i, MKT_COLS, row, alt=(i % 2 == 1))
    ws.freeze_panes = "A6"
    print(f"  Solapa {tab_name} creada ({len(rows)} registros)")


def build_markets(wb, fecha_dt, hora_dt):
    print(f"  Consultando {MARKETS_URL}...")
    try:
        r = requests.get(MARKETS_URL, timeout=30); r.raise_for_status()
        payload = r.json()
    except Exception as e:
        print(f"  [ERROR] No se pudo obtener datos de mercados: {e}"); return

    try:
        indices    = payload.get("world_indices", [])
        assets     = payload.get("assets", {})
        commodities = assets.get("commodities", [])
        currencies  = assets.get("currencies", [])
        ust         = assets.get("ust", [])
        crypto      = payload.get("crypto", [])

        us_rows   = [_parse_mkt_row(d) for d in indices if d.get("symbol") in _US_SYMBOLS]
        am_rows   = [_parse_mkt_row(d) for d in indices if d.get("symbol") in _AM_SYMBOLS]
        eu_rows   = [_parse_mkt_row(d) for d in indices if d.get("symbol") in _EU_SYMBOLS]
        asia_rows = [_parse_mkt_row(d) for d in indices if d.get("symbol") in _ASIA_SYMBOLS]

        _build_mkt_sheet(wb, "INDICES US",   TAB_IDX_US,   "Índices EE.UU. — VIX / S&P 500 / Dow / Nasdaq",    us_rows,                  fecha_dt, hora_dt)
        _build_mkt_sheet(wb, "INDICES AM",   TAB_IDX_AM,   "Índices Américas — IBOVESPA / MERVAL / TSX",        am_rows,                  fecha_dt, hora_dt)
        _build_mkt_sheet(wb, "INDICES EU",   TAB_IDX_EU,   "Índices Europa — FTSE / DAX / CAC 40 / IBEX",       eu_rows,                  fecha_dt, hora_dt)
        _build_mkt_sheet(wb, "INDICES ASIA", TAB_IDX_ASIA, "Índices Asia — Nikkei / Hang Seng / KOSPI / ASX",   asia_rows,                fecha_dt, hora_dt)
        _build_mkt_sheet(wb, "MONEDAS",      TAB_MONEDAS,  "Monedas — Tipos de Cambio",                         [_parse_mkt_row(d) for d in currencies],  fecha_dt, hora_dt)
        _build_mkt_sheet(wb, "COMMODITIES",  TAB_CMDTY,    "Commodities — Metales / Energía",                   [_parse_mkt_row(d) for d in commodities], fecha_dt, hora_dt)
        _build_mkt_sheet(wb, "US BONDS",     TAB_BONDS,    "US Treasury Bonds — Tasas / Yields",                [_parse_mkt_row(d) for d in ust],         fecha_dt, hora_dt)
        _build_mkt_sheet(wb, "CRIPTO",       TAB_CRIPTO,   "Criptomonedas (USD)",                               [_parse_mkt_row(d) for d in crypto],      fecha_dt, hora_dt)

    except Exception as e:
        import traceback
        print(f"  [ERROR] Falló build_markets: {e}")
        traceback.print_exc()


# ═══════════════════════════════════════════════════════════════════════════════
# VARIABLES MONETARIAS
# ═══════════════════════════════════════════════════════════════════════════════
VARMON_BASE = "https://api.alquimiaconsultora.com/output/datosmacro"

def _varmon_get(slug):
    try:
        r = requests.get(f"{VARMON_BASE}/{slug}.json", timeout=30); r.raise_for_status()
        data = r.json()
        return data if isinstance(data, list) else []
    except Exception as e:
        print(f"  [ERROR] GET {slug}: {e}"); return []

def _last_row(rows):
    valid = [r for r in rows if r.get("Date")]
    if not valid: return None
    return max(valid, key=lambda r: str(r.get("Date", "")))

def _last_value(rows, cat1, cat2, cat3, cat4):
    filtered = [
        r for r in rows
        if str(r.get("Category 1", "")).strip() == cat1
        and str(r.get("Category 2", "")).strip() == cat2
        and str(r.get("Category 3", "")).strip() == cat3
        and str(r.get("Category 4", "")).strip() == cat4
    ]
    if not filtered: return None, None
    last = max(filtered, key=lambda r: str(r.get("Date", "")))
    return last.get("Value"), last.get("Date")

def _write_section_title(ws, row, title, ncols=4):
    c = ws.cell(row, 1)
    c.value = title
    c.font  = _font(FG_TITLE, bold=True, size=9)
    c.fill  = _fill(BG_TITLE)
    ws.row_dimensions[row].height = 16
    for col in range(2, ncols + 1):
        ws.cell(row, col).fill = _fill(BG_TITLE)
    return row + 1

def _write_varmon_hdr(ws, row):
    labels = ["Variable", "Fecha", "Valor", "Unidad"]
    widths  = [40, 12, 16, 20]
    ws.row_dimensions[row].height = 16
    for i, (lbl, w) in enumerate(zip(labels, widths), 1):
        c = ws.cell(row, i)
        c.value = lbl; c.fill = _fill(BG_HDR)
        c.font  = _font(FG_WHITE, bold=True)
        c.alignment = _center(); c.border = _border()
        ws.column_dimensions[get_column_letter(i)].width = w
    return row + 1

def _write_varmon_row(ws, row, label, fecha, valor, unidad, alt=False):
    bg = BG_ALT if alt else BG_ROW
    ws.row_dimensions[row].height = 15
    data = [label, fecha, valor, unidad]
    for i, val in enumerate(data, 1):
        c = ws.cell(row, i)
        c.value = val; c.fill = _fill(bg)
        c.font  = _font(FG_DATA); c.border = _border()
        if i == 3 and val is not None:
            c.alignment = _right()
            if unidad == "Billones ARS":
                c.number_format = "#,##0.0"
            elif unidad == "MM USD CCL":
                c.number_format = "#,##0.000"
            elif unidad == "% TNA":
                c.number_format = "0.00"
            else:
                c.number_format = "#,##0.00"
        else:
            c.alignment = _center() if i == 2 else _left() if i in (1,4) else _right()
    return row + 1

def _get_ccl_al30_24hs():
    try:
        r = requests.get("https://api.alquimiaconsultora.com/output/dolares.json", timeout=30)
        r.raise_for_status()
        data = r.json()
        ccl = data.get("data", {}).get("ccl", {}).get("al30", {}).get("24hs", {}).get("price")
        return _f(ccl)
    except Exception as e:
        print(f"  [WARN] No se pudo obtener CCL: {e}")
        return None


def _get_bcra_serie_last(serie_key):
    """Obtiene el último valor de una serie de bcra.json (estructura: {series: {KEY: {fecha: valor}}})."""
    try:
        r = requests.get(BCRA_URL, timeout=30); r.raise_for_status()
        data = r.json()
        serie = data.get("series", {}).get(serie_key)
        if not serie or not isinstance(serie, dict):
            print(f"  [WARN] Serie '{serie_key}' no encontrada en bcra.json")
            return None, None
        ultima_fecha = max(serie.keys())
        return _f(serie[ultima_fecha]), ultima_fecha
    except Exception as e:
        print(f"  [WARN] Error obteniendo '{serie_key}' de bcra.json: {e}")
        return None, None

def build_varmon(wb, fecha_dt, hora_dt):
    print("  Consultando monetary-aggregates...")
    agg  = _varmon_get("monetary-aggregates")
    print("  Consultando market-interest-rates...")
    mirt = _varmon_get("market-interest-rates")
    print("  Obteniendo CCL AL30 24hs...")
    ccl  = _get_ccl_al30_24hs()
    print(f"  CCL AL30 24hs: {ccl}")
    print("  Obteniendo TAMAR y BADLAR desde bcra.json...")
    tamar_valor, tamar_fecha   = _get_bcra_serie_last("TAMAR")
    badlar_valor, badlar_fecha = _get_bcra_serie_last("BADLAR")
    print(f"  TAMAR: {tamar_fecha} → {tamar_valor}  |  BADLAR: {badlar_fecha} → {badlar_valor}")

    ws = wb.create_sheet("VAR. MONETARIAS")
    ws.sheet_properties.tabColor = TAB_VARMON
    write_meta(ws, "Variables Monetarias — Últimos Datos Disponibles", fecha_dt, hora_dt, None)

    row = 5

    last_agg = _last_row(agg)
    fecha_agg = last_agg.get("Date", "") if last_agg else "—"

    items_agg = [
        ("M0 — Base Monetaria", last_agg.get("Monetary Base") if last_agg else None),
        ("M1 Total",            last_agg.get("Total M1")      if last_agg else None),
        ("M2 Total",            last_agg.get("Total M2")      if last_agg else None),
        ("M3 Total",            last_agg.get("Total M3")      if last_agg else None),
    ]

    row = _write_section_title(ws, row, "Agregados Monetarios (Billones de ARS)")
    row = _write_varmon_hdr(ws, row)
    for i, (label, valor) in enumerate(items_agg):
        val_billones = round(valor / 1_000, 1) if valor is not None else None
        row = _write_varmon_row(ws, row, label, fecha_agg, val_billones, "Billones ARS", alt=(i % 2 == 1))

    row += 1

    row = _write_section_title(ws, row, f"Agregados Monetarios (USD CCL — AL30 24hs: {ccl:,.2f})" if ccl else "Agregados Monetarios (USD CCL — sin dato)")
    row = _write_varmon_hdr(ws, row)
    for i, (label, valor) in enumerate(items_agg):
        if valor is not None and ccl and ccl > 0:
            val_usd = round((valor / ccl) / 1_000, 3)
        else:
            val_usd = None
        row = _write_varmon_row(ws, row, label, fecha_agg, val_usd, "MM USD CCL", alt=(i % 2 == 1))

    row += 1

    row = _write_section_title(ws, row, "Tasas de Mercado — TNA (últimos datos disponibles)")
    row = _write_varmon_hdr(ws, row)

    # TAMAR y BADLAR desde bcra.json; el resto desde market-interest-rates
    tasas_bcra = [
        ("TAMAR — Bancos Privados",   tamar_valor,  tamar_fecha),
        ("BADLAR — Bancos Privados",  badlar_valor, badlar_fecha),
    ]
    tasas_mirt = [
        ("TM20 — Total",              "TM20",           "Pesos",          "Total", "TNA"),
        ("Call en Pesos — Total",     "Call en pesos",  "Total",          "Tasa",  "TNA"),
        ("Plazo Fijo Pesos — Total",  "Plazo Fijo",     "Pesos",          "Total", "TNA"),
        ("Pases 1 día",               "Pases",          "Entre terceros", "1 día", "TNA"),
    ]

    fila_idx = 0
    for label, valor, fecha in tasas_bcra:
        row = _write_varmon_row(ws, row, label, fecha or "—", valor, "% TNA", alt=(fila_idx % 2 == 1))
        fila_idx += 1
    for label, c1, c2, c3, c4 in tasas_mirt:
        valor, fecha = _last_value(mirt, c1, c2, c3, c4)
        row = _write_varmon_row(ws, row, label, fecha or "—", valor, "% TNA", alt=(fila_idx % 2 == 1))
        fila_idx += 1

    row += 1

    row = _write_section_title(ws, row, "Reservas Internacionales (USD MM)")
    row = _write_varmon_hdr(ws, row)
    try:
        r_res = requests.get(f"{VARMON_BASE}/daily-monetary-report.json", timeout=30)
        r_res.raise_for_status()
        daily = r_res.json()
        if isinstance(daily, list):
            reservas_rows = [
                r for r in daily
                if any("reserva" in str(v).lower() for v in r.values()
                       if isinstance(v, str))
            ]
            if reservas_rows:
                ultimo_res = max(reservas_rows, key=lambda r: str(r.get("Date", "")))
                res_fecha  = str(ultimo_res.get("Date", "—"))[:10]
                res_valor  = _f(ultimo_res.get("Value"))
            else:
                res_valor, res_fecha = None, "—"
        else:
            res_valor, res_fecha = None, "—"
    except Exception as e:
        print(f"  [WARN] No se pudo obtener reservas: {e}")
        res_valor, res_fecha = None, "—"

    row = _write_varmon_row(ws, row, "Reservas Internacionales", res_fecha, res_valor, "USD MM", alt=False)

    ws.freeze_panes = "A6"
    print(f"  Solapa VAR. MONETARIAS creada ({row - 5} filas)")


# ═══════════════════════════════════════════════════════════════════════════════
# ROFEX — DÓLAR FUTURO
# ═══════════════════════════════════════════════════════════════════════════════
ROFEX_URL = "https://api.alquimiaconsultora.com/output/futuros.json"

ROFEX_COLS = [
    {"key": "mes",         "label": "Contrato",    "w": 11},
    {"key": "vencimiento", "label": "Vencimiento", "w": 14},
    {"key": "days_to_mat", "label": "Días",        "w":  7, "align": "right"},
    {"key": "rofex",       "label": "ROFEX",       "w": 11, "align": "right", "fmt": "#,##0.00"},
    {"key": "tna",         "label": "TNA %",       "w": 10, "align": "right", "fmt": "0.00"},
    {"key": "tir",         "label": "TIR %",       "w": 10, "align": "right", "fmt": "0.00"},
    {"key": "pase",        "label": "Pase (ARS)",  "w": 12, "align": "right", "fmt": "#,##0.00"},
]

def build_rofex(wb, fecha_dt, hora_dt):
    print(f"  Consultando {ROFEX_URL}...")
    try:
        r = requests.get(ROFEX_URL, timeout=30); r.raise_for_status()
        payload = r.json()
    except Exception as e:
        print(f"  [ERROR] No se pudo obtener datos de ROFEX: {e}"); return

    vencimientos = payload.get("vencimientos", [])
    if not vencimientos:
        print("  [WARN] Sin datos de vencimientos, se omite solapa ROFEX."); return

    spot = None
    curva = payload.get("curva", [])
    if curva:
        spot_entry = next((x for x in curva if x.get("mes", "").startswith("SPOT")), None)
        if spot_entry:
            spot = spot_entry.get("rofex")

    ws = wb.create_sheet("ROFEX")
    ws.sheet_properties.tabColor = TAB_ROFEX

    try:
        titulo = "Dólar Futuro ROFEX — Curva de Tasas"
        write_meta(ws, titulo, fecha_dt, hora_dt, None)

        r5 = ws.cell(5, 1)
        spot_txt = f"Spot (ref): {spot:,.2f}" if spot is not None else "Curva de Futuros"
        r5.value = spot_txt
        r5.font = _font(FG_TITLE, bold=True, size=9); r5.fill = _fill(BG_TITLE)
        ws.row_dimensions[5].height = 14

        write_hdr(ws, 6, ROFEX_COLS)
        row_idx = 7

        for i, item in enumerate(vencimientos):
            fila = {
                "mes":         item.get("mes"),
                "vencimiento": item.get("vencimiento"),
                "days_to_mat": item.get("days_to_mat"),
                "rofex":       _f(item.get("rofex")),
                "tna":         _f(item.get("tna")),
                "tir":         _f(item.get("tir")),
                "pase":        _f(item.get("pase")),
            }
            write_row(ws, row_idx, ROFEX_COLS, fila, alt=(row_idx % 2 == 0))
            row_idx += 1

        ws.freeze_panes = "A7"
        print(f"  Solapa ROFEX creada ({len(vencimientos)} contratos)")
    except Exception as e:
        import traceback
        print(f"  [ERROR] Falló al construir solapa ROFEX: {e}")
        traceback.print_exc()


# ═══════════════════════════════════════════════════════════════════════════════
# CAUCIONES
# ═══════════════════════════════════════════════════════════════════════════════
CAUCIONES_URL = "https://api.alquimiaconsultora.com/output/cauciones.json"

CAUCIONES_COLS = [
    {"key": "plazoDias", "label": "Plazo (días)", "w": 13, "align": "right"},
    {"key": "tna",       "label": "TNA %",        "w": 12, "align": "right", "fmt": "0.00%"},
]

def build_cauciones(wb, fecha_dt, hora_dt):
    print(f"  Consultando {CAUCIONES_URL}...")
    try:
        r = requests.get(CAUCIONES_URL, timeout=30); r.raise_for_status()
        payload = r.json()
    except Exception as e:
        print(f"  [ERROR] No se pudo obtener datos de cauciones: {e}"); return

    quotes = payload.get("quotes", [])
    if not quotes:
        print("  [WARN] Sin datos de cauciones, se omite solapa."); return

    mercado = payload.get("mercado", "BYMA")
    moneda  = payload.get("moneda", "ARS")

    ws = wb.create_sheet("CAUCIONES")
    ws.sheet_properties.tabColor = TAB_CAUCIONES

    try:
        write_meta(ws, f"Cauciones — {mercado} ({moneda})", fecha_dt, hora_dt, None)

        r5 = ws.cell(5, 1)
        r5.value = f"Mercado: {mercado}  |  Moneda: {moneda}"
        r5.font = _font(FG_TITLE, bold=True, size=9); r5.fill = _fill(BG_TITLE)
        ws.row_dimensions[5].height = 14

        write_hdr(ws, 6, CAUCIONES_COLS)
        row_idx = 7

        for item in quotes:
            fila = {
                "plazoDias": item.get("plazoDias"),
                "tna":       _f(item.get("tna")),
            }
            write_row(ws, row_idx, CAUCIONES_COLS, fila, alt=(row_idx % 2 == 0))
            row_idx += 1

        ws.freeze_panes = "A7"
        print(f"  Solapa CAUCIONES creada ({len(quotes)} plazos)")
    except Exception as e:
        import traceback
        print(f"  [ERROR] Falló al construir solapa CAUCIONES: {e}")
        traceback.print_exc()


# ═══════════════════════════════════════════════════════════════════════════════
# UPLOAD
# ═══════════════════════════════════════════════════════════════════════════════
def _get_auth_headers() -> dict:
    secret = os.getenv("EXPORT_SECRET", "alquimia-export-2026")
    return {"X-Export-Secret": secret}


def write_export_meta(fname: str, drive_ok: bool, drive_url: str | None,
                      drive_backup_ok: bool, sections: list, errors: list):
    from datetime import timezone
    meta = {
        "updatedAtUtc": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "file": fname,
        "driveOk": drive_ok,
        "driveUrl": drive_url,
        "driveBackupOk": drive_backup_ok,
        "sections": sections,
        "errors": errors,
    }
    try:
        EXPORT_META_JSON.parent.mkdir(parents=True, exist_ok=True)
        tmp = EXPORT_META_JSON.with_suffix(".tmp")
        tmp.write_text(json.dumps(meta, ensure_ascii=False), encoding="utf-8")
        tmp.replace(EXPORT_META_JSON)
        print(f"  [META] export_meta.json escrito → {EXPORT_META_JSON}")
    except Exception as e:
        print(f"  [META] Error escribiendo export_meta.json: {e}")


def drive_upload(fpath: Path, endpoint: str) -> str | None:
    headers = _get_auth_headers()
    try:
        with open(fpath,"rb") as f:
            resp = requests.post(
                f"{API_BASE}{endpoint}",
                headers=headers,
                files={"file": (fpath.name, f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                timeout=60)
        if resp.status_code == 200:
            data = resp.json()
            action = "Actualizado" if data.get("updated") else "Creado"
            print(f"  [DRIVE] {action} → {data.get('url') or data.get('fileId')}")
            return data.get("fileId")
        elif resp.status_code == 404:
            print(f"  [DRIVE] {endpoint} no disponible (404) — pendiente de implementar en backend")
            return None
        else:
            print(f"  [DRIVE] Error HTTP {resp.status_code}: {resp.text[:200]}")
            return None
    except Exception as e:
        print(f"  [DRIVE] Error: {e}"); return None


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    global API_BASE, OUTPUT_DIR
    ap = argparse.ArgumentParser()
    ap.add_argument("--api",     default=API_BASE)
    ap.add_argument("--output",  default=str(OUTPUT_DIR))
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    API_BASE   = args.api.rstrip("/")
    OUTPUT_DIR = Path(args.output)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    now       = datetime.now(TZ_AR)
    fecha_hoy = now.strftime("%Y-%m-%d")
    fecha_dt  = now.date()
    hora_dt   = now.time()
    fecha_str = now.strftime("%d/%m/%Y")
    hora_str  = now.strftime("%H:%M")

    print(f"\n{'='*62}")
    print(f"  EXPORT CIERRE  v3.0  |  {fecha_str} {hora_str} ART")
    print(f"  API:     {API_BASE}")
    print(f"  OUTPUT:  {OUTPUT_DIR}")
    print(f"{'='*62}")

    if args.dry_run:
        for path in ["/api/lecaps","/api/boncer","/api/dlk","/api/tamar",
                     "/api/duales","/api/onln","/api/onlny","/api/bonoshd","/api/bote"]:
            try:
                r = requests.get(f"{API_BASE}{path}", timeout=10); d = r.json()
                rows = d.get("rows",d) if isinstance(d,dict) else d
                print(f"  ✓  {path:<18} HTTP {r.status_code}  |  {len(rows) if isinstance(rows,list) else 0} items")
            except Exception as e:
                print(f"  ✗  {path:<18} {e}")
        return

    # Cargar historia
    ultimo = OUTPUT_DIR / "ultimo_cierre.xlsx"
    print(f"\n[HIST] Cargando desde {ultimo}...")
    hist = load_historia_from_xlsx(ultimo)
    total_entradas = sum(len(v) for v in hist.values())
    print(f"  {total_entradas} entradas previas ({len(hist)} instrumentos)")
    for instr, entries in sorted(hist.items()):
        fechas = sorted(set(e.get("fecha","") for e in entries))
        print(f"    {instr}: {len(entries)} snapshots | fechas: {', '.join(fechas[-5:])}")

    # Construir workbook
    wb = Workbook(); wb.remove(wb.active)

    print("\n[1/9] LECAPS");       build_lecaps(wb, hist, fecha_hoy, hora_str, fecha_dt, hora_dt)
    print("\n[2/9] BONCER");       build_boncer(wb, hist, fecha_hoy, hora_str, fecha_dt, hora_dt)
    print("\n[3/9] DLK");          build_dlk(wb, hist, fecha_hoy, hora_str, fecha_dt, hora_dt)
    print("\n[4/9] TAMAR");        build_tamar(wb, hist, fecha_hoy, hora_str, fecha_dt, hora_dt)
    print("\n[5/9] DUALES");       build_duales(wb, hist, fecha_hoy, hora_str, fecha_dt, hora_dt)
    print("\n[6/9] ONs LN");       build_ons(wb, hist, "/api/onln",  "ONS_LN",  "ONs_LN",
                                             "ONs Ley Nacional — USD MEP / CCL",  TAB_ONLN,
                                             fecha_hoy, hora_str, fecha_dt, hora_dt)
    print("\n[7/9] ONs LNY");      build_ons(wb, hist, "/api/onlny", "ONS_LNY", "ONs_LNY",
                                             "ONs Ley Nueva York — USD MEP / CCL", TAB_ONLNY,
                                             fecha_hoy, hora_str, fecha_dt, hora_dt)
    print("\n[8/9] HD Soberanos"); build_hdsob(wb, hist, fecha_hoy, hora_str, fecha_dt, hora_dt)
    print("\n[9/9] BOTE");          build_bote(wb, hist, fecha_hoy, hora_str, fecha_dt, hora_dt)
    print("\n[+] HISTORIA");        build_historia(wb, hist)
    print("\n[+] DÓLAR");           build_dolar(wb, fecha_dt, hora_dt)
    print("\n[+] ACCIONES AR");     build_acciones(wb, fecha_dt, hora_dt)
    print("\n[+] ADR");             build_adr(wb, fecha_dt, hora_dt)
    print("\n[+] NYSE");             build_nyse(wb, fecha_dt, hora_dt)
    print("\n[+] CEDEAR");          build_cedear(wb, fecha_dt, hora_dt)
    print("\n[+] MERCADOS");         build_markets(wb, fecha_dt, hora_dt)
    print("\n[+] VAR. MONETARIAS");   build_varmon(wb, fecha_dt, hora_dt)
    print("\n[+] ROFEX");             build_rofex(wb, fecha_dt, hora_dt)
    print("\n[+] CAUCIONES");         build_cauciones(wb, fecha_dt, hora_dt)

    if not wb.sheetnames:
        print("\n[ERROR] Sin datos."); sys.exit(1)

    # Guardar
    fname = f"cierre_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    fpath = OUTPUT_DIR / fname
    wb.save(str(fpath))
    shutil.copy(str(fpath), str(ultimo))

    print(f"\n{'='*62}")
    print(f"  ✅  {fpath.name}")
    print(f"  ✅  ultimo_cierre.xlsx")

    print("\n[DRIVE] Subiendo ultimo_cierre.xlsx...")
    drive_url    = drive_upload(ultimo, "/api/admin/upload-cierre")
    drive_ok     = drive_url is not None

    print("[DRIVE] Subiendo backup diario a diarios/...")
    backup_id    = drive_upload(fpath, "/api/admin/upload-cierre-diario")
    backup_ok    = backup_id is not None

    write_export_meta(
        fname         = fname,
        drive_ok      = drive_ok,
        drive_url     = drive_url,
        drive_backup_ok = backup_ok,
        sections      = ["LECAPS","BONCER","DLK","TAMAR","DUALES","ONs LN","ONs LNY","BOTE",
                         "DÓLAR","ACCIONES AR","ADR","NYSE","CEDEAR","MERCADOS","VAR. MONETARIAS"],
        errors        = [],
    )

    try:
        with open(EXPORT_META_JSON, "rb") as f_meta:
            resp = requests.post(
                f"{API_BASE}/upload",
                headers={"X-Upload-Token": os.getenv("UPLOAD_TOKEN", "alquimia2025")},
                files={"file": ("export_meta.json", f_meta, "application/json")},
                params={"name": "export_meta.json"},
                timeout=30,
            )
        print(f"  [META] export_meta.json subido → {resp.status_code}")
    except Exception as e:
        print(f"  [META] Error subiendo export_meta.json: {e}")

    print(f"\n{'='*62}\n")


if __name__ == "__main__":
    main()