"""
export_cierre.py  v3.0
======================
Cambios v3.0:
    - HISTORIA: columna Hora agregada
    - HISTORIA: columnas con nombres reales por instrumento (TEM/TNA/TIREA/MD/WAL etc.)
    - Fix _d(): soporta formato DD/MM/YYYY (BONCER fecha vencimiento)
    - Backup diario sube a Drive en subcarpeta diarios/ via /api/admin/upload-cierre-diario
    - MD siempre con 2 decimales
"""

import os, sys, shutil, argparse
from datetime import datetime, date, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import google.auth.transport.requests as google_requests
    from google.oauth2 import service_account
    GOOGLE_AUTH_AVAILABLE = True
except ImportError:
    GOOGLE_AUTH_AVAILABLE = False

SERVICE_ACCOUNT_FILE = Path(__file__).parent / "service_account.json" 

# ─── Config ───────────────────────────────────────────────────────────────────
TZ_AR      = ZoneInfo("America/Argentina/Buenos_Aires")
API_BASE   = os.getenv("API_BASE_URL", "https://api.alquimiaconsultora.com").rstrip("/")
OUTPUT_DIR = Path(os.getenv("OUTPUT_DIR", "./output/reportes"))

# ─── Paleta ───────────────────────────────────────────────────────────────────
BG_TITLE = "0A1520"; BG_HDR = "1F3864"; BG_ROW = "0D1B2A"; BG_ALT = "111E2D"
FG_WHITE = "FFFFFF"; FG_DATA = "D8E4F0"; FG_MUTED = "7A9BBB"
FG_TITLE = "00AEEF"; FG_BORD = "1E3355"

TAB_LECAPS = "00AEEF"; TAB_BONCER = "9B59B6"; TAB_DLK    = "E67E22"
TAB_TAMAR  = "27AE60"; TAB_DUAL   = "F39C12"; TAB_ONLN   = "4CAF50"
TAB_ONLNY  = "FF9800"; TAB_HDSOB  = "E74C3C"; TAB_BOTE   = "16A085"
TAB_HIST   = "888888"; TAB_DOLAR  = "2ECC71"; TAB_ADR = "3498DB"; TAB_CEDEAR = "E91E8C"; TAB_NYSE = "F39C12"

# ─── Estilo ───────────────────────────────────────────────────────────────────
def _fill(c): return PatternFill("solid", start_color=c, fgColor=c)
def _font(c=FG_DATA, bold=False, size=9): return Font(name="Arial", color=c, bold=bold, size=size)
def _center(): return Alignment(horizontal="center", vertical="center")
def _right():  return Alignment(horizontal="right",  vertical="center")
def _border():
    s = Side(style="thin", color=FG_BORD)
    return Border(left=s, right=s, top=s, bottom=s)

def paint_bg(ws, rows=300, cols=40):
    for r in range(1, rows+1):
        for c in range(1, cols+1):
            ws.cell(r, c).fill = _fill(BG_TITLE)

def write_meta(ws, titulo, fecha_dt, hora_dt, liq_dt):
    paint_bg(ws)
    ws.sheet_view.showGridLines = False
    ws["A1"].value = titulo
    ws["A1"].font  = Font(name="Arial", bold=True, color=FG_TITLE, size=13)
    ws["A1"].fill  = _fill(BG_TITLE)
    ws.row_dimensions[1].height = 22

    ws["A2"].value = "Precios al:"; ws["A2"].font = _font(FG_MUTED, size=8); ws["A2"].fill = _fill(BG_TITLE)
    ws["B2"].value = fecha_dt;      ws["B2"].number_format = "DD/MM/YYYY"
    ws["B2"].font  = _font(FG_DATA, bold=True, size=8); ws["B2"].fill = _fill(BG_TITLE)
    ws["C2"].value = hora_dt;       ws["C2"].number_format = "HH:MM"
    ws["C2"].font  = _font(FG_DATA, size=8); ws["C2"].fill = _fill(BG_TITLE)
    ws["D2"].value = "ART"; ws["D2"].font = _font(FG_MUTED, size=8); ws["D2"].fill = _fill(BG_TITLE)

    ws["A3"].value = "Liquidación:"; ws["A3"].font = _font(FG_MUTED, size=8); ws["A3"].fill = _fill(BG_TITLE)
    if liq_dt:
        ws["B3"].value = liq_dt; ws["B3"].number_format = "DD/MM/YYYY"
        ws["B3"].font  = _font(FG_DATA, size=8); ws["B3"].fill = _fill(BG_TITLE)

    ws.column_dimensions["T"].width = 14
    c = ws.cell(1, 20); c.value = "ALQUIMIA"
    c.font = Font(name="Arial", bold=True, color=FG_TITLE, size=11)
    c.fill = _fill(BG_TITLE); c.alignment = Alignment(horizontal="right", vertical="center")
    c2 = ws.cell(2, 20); c2.value = "Consultora"
    c2.font = _font(FG_MUTED, size=8); c2.fill = _fill(BG_TITLE)
    c2.alignment = Alignment(horizontal="right", vertical="center")

def write_hdr(ws, row, cols):
    ws.row_dimensions[row].height = 16
    for i, col in enumerate(cols, 1):
        c = ws.cell(row, i)
        c.value = col["label"]; c.fill = _fill(BG_HDR)
        c.font = _font(FG_WHITE, bold=True); c.alignment = _center(); c.border = _border()
        ws.column_dimensions[get_column_letter(i)].width = col.get("w", 11)

def write_row(ws, row, cols, data, alt=False):
    ws.row_dimensions[row].height = 15
    bg = BG_ALT if alt else BG_ROW
    for i, col in enumerate(cols, 1):
        c = ws.cell(row, i)
        c.value = data.get(col["key"]); c.fill = _fill(bg); c.font = _font(FG_DATA)
        c.alignment = _right() if col.get("align") == "right" else _center()
        c.border = _border()
        if col.get("fmt"): c.number_format = col["fmt"]

# ─── Utilidades ───────────────────────────────────────────────────────────────
def _f(v):
    try: return float(v) if v is not None else None
    except: return None

def _d(v):
    """Parsea fecha. Soporta YYYY-MM-DD y DD/MM/YYYY."""
    if v is None: return None
    s = str(v).strip()[:10]
    if len(s) == 10 and s[2] == "/" and s[5] == "/":
        try: return datetime.strptime(s, "%d/%m/%Y").date()
        except: return None
    try: return date.fromisoformat(s)
    except: return None

def _pct(v):
    x = _f(v)
    return None if x is None else x / 100.0

def api_get(path):
    try:
        r = requests.get(f"{API_BASE}{path}", timeout=30); r.raise_for_status()
        d = r.json(); return d if isinstance(d, list) else []
    except Exception as e:
        print(f"  [ERROR] GET {path}: {e}"); return []

def api_get_obj(path):
    try:
        r = requests.get(f"{API_BASE}{path}", timeout=30); r.raise_for_status()
        return r.json()
    except Exception as e:
        print(f"  [ERROR] GET {path}: {e}"); return {}

def _filtro_comun(precio, volumen, tir, vto, hoy):
    if vto and vto <= hoy:   return False
    if not precio or precio <= 0:   return False
    if not volumen or volumen <= 0: return False
    if tir is not None and (tir < -0.5 or tir > 5.0): return False
    return True

# ─── Historia ─────────────────────────────────────────────────────────────────
HIST_MAP = {
    "LECAPS": dict(dim="dias",     precio="precio",    var="var",  vol="volumen", tem="tem",  tna="tna",       tirea="tirea",       extra=None,        extra_lbl="",           dim_fmt="#,##0",  precio_fmt="#,##0.000", extra_fmt=""),
    "BONCER": dict(dim="md",       precio="precio",    var="var",  vol="volumen", tem=None,   tna="tna",       tirea="tirea",       extra=None,        extra_lbl="",           dim_fmt="0.00",   precio_fmt="#,##0.00",  extra_fmt=""),
    "DLK":    dict(dim="md",       precio="precio_ars",var="var",  vol="volumen", tem=None,   tna="tna",       tirea="tirea",       extra=None,        extra_lbl="",           dim_fmt="0.00",   precio_fmt="#,##0.00",  extra_fmt=""),
    "TAMAR":  dict(dim="dias_vto", precio="precio",    var="var",  vol="volumen", tem=None,   tna="tna_tamar", tirea="tirea_tamar", extra=None,        extra_lbl="",           dim_fmt="#,##0",  precio_fmt="#,##0.00",  extra_fmt=""),
    "DUAL":   dict(dim="dias_vto", precio="precio",    var="var",  vol="volumen", tem=None,   tna=None,        tirea="tirea_tamar", extra="tirea_fija",extra_lbl="TIREA Fija", dim_fmt="#,##0",  precio_fmt="#,##0.00",  extra_fmt="0.00%"),
    "ONS_LN": dict(dim="md",       precio="precio",    var="var",  vol="volumen", tem=None,   tna=None,        tirea="tirea",       extra="tirea_ccl", extra_lbl="TIREA CCL",  dim_fmt="0.00",   precio_fmt="#,##0.0000",extra_fmt="0.00%"),
    "ONS_LNY":dict(dim="md",       precio="precio",    var="var",  vol="volumen", tem=None,   tna=None,        tirea="tirea",       extra="tirea_ccl", extra_lbl="TIREA CCL",  dim_fmt="0.00",   precio_fmt="#,##0.0000",extra_fmt="0.00%"),
    "HDSOB":  dict(dim="md",       precio="precio",    var="var",  vol="volumen", tem=None,   tna="tna",       tirea="tirea",       extra=None,        extra_lbl="",           dim_fmt="0.00",   precio_fmt="#,##0.00",  extra_fmt=""),
    "BOTE":   dict(dim="md",       precio="precio_sucio",var="var",vol="volumen", tem=None,   tna="tna",       tirea="tirea",       extra="wal",       extra_lbl="WAL",        dim_fmt="0.00",   precio_fmt="#,##0.00",  extra_fmt="0.00"),
}
HIST_FIXED_HDRS   = ["Precio", "Var%",  "TEM",  "TNA",  "TIREA", "Días/MD", "Volumen", "Extra"]
HIST_FIXED_WIDTHS = [12,        8,       10,     10,     10,       9,          14,        13]


def load_historia_from_xlsx(ultimo_path: Path) -> dict:
    """
    Lee la hoja HISTORIA del Excel existente y reconstruye el diccionario hist
    en formato snap moderno (igual al que produce add_snapshot_to_hist).
    Columnas esperadas (v3): Instrumento | Fecha | Hora | Ticker |
                             Precio | Var% | TEM | TNA | TIREA | Días/MD | Volumen | Extra
    """
    hist = {}
    if not ultimo_path.exists(): return hist
    try:
        wb = load_workbook(str(ultimo_path), read_only=True, data_only=True)
        if "HISTORIA" not in wb.sheetnames:
            wb.close(); return hist
        ws = wb["HISTORIA"]

        # Detectar si tiene columna Hora (v3) o no (v2)
        hdr = None
        for row in ws.iter_rows(min_row=3, max_row=3, values_only=True):
            hdr = row; break
        has_hora = hdr and len(hdr) > 2 and str(hdr[2] or "").upper() == "HORA"

        # Agrupar filas por (instr, fecha, hora)
        current = {}  # key: (instr, fecha, hora) -> list of snap dicts
        for row in ws.iter_rows(min_row=4, values_only=True):
            instr = row[0]
            if not instr or str(instr) == "Instrumento": continue
            instr = str(instr).upper()
            fecha = str(row[1])[:10] if row[1] else None
            if not fecha: continue

            if has_hora:
                hora_raw = row[2]
                if hasattr(hora_raw, "strftime"):
                    hora = hora_raw.strftime("%H:%M")
                elif hora_raw is not None:
                    hora = str(hora_raw)[:5]
                else:
                    hora = ""
                ticker  = str(row[3] or "")
                v_precio= row[4]
                v_var   = row[5]
                v_tem   = row[6]
                v_tna   = row[7]
                v_tirea = row[8]
                v_dim   = row[9]
                v_vol   = row[10]
                v_extra = row[11] if len(row) > 11 else None
            else:
                hora    = ""
                ticker  = str(row[2] or "")
                v_precio= None  # versión vieja no tenía precio
                v_var   = None
                v_tem   = row[3]
                v_tna   = row[4]
                v_tirea = row[5]
                v_dim   = row[6] if len(row) > 6 else None
                v_vol   = row[7] if len(row) > 7 else None
                v_extra = row[8] if len(row) > 8 else None

            m = HIST_MAP.get(instr)
            if not m: continue

            # Reconstruir el snap con las claves que espera build_historia
            snap_row = {"ticker": ticker}
            if m["dim"]:    snap_row[m["dim"]]    = v_dim
            if m["precio"]: snap_row[m["precio"]] = v_precio
            if m["var"]:    snap_row[m["var"]]    = v_var
            if m["vol"]:    snap_row[m["vol"]]    = v_vol
            if m["tem"]:    snap_row[m["tem"]]    = v_tem
            if m["tna"]:    snap_row[m["tna"]]    = v_tna
            if m["tirea"]:  snap_row[m["tirea"]]  = v_tirea
            if m["extra"]:  snap_row[m["extra"]]  = v_extra

            key = (instr, fecha, hora)
            if key not in current: current[key] = []
            current[key].append(snap_row)

        # Convertir al formato hist estándar
        for (instr, fecha, hora), snap_rows in current.items():
            if instr not in hist: hist[instr] = []
            hist[instr].append({"fecha": fecha, "hora": hora, "snap": snap_rows})

        # Ordenar y limitar a 365 días por instrumento
        for instr in hist:
            hist[instr] = sorted(hist[instr], key=lambda h: h.get("fecha", ""))[-365:]

        wb.close()
    except Exception as e:
        print(f"  [HIST] Error leyendo historia: {e}")
    return hist


def add_snapshot_to_hist(hist, instr, fecha, hora, snap_rows):
    if instr not in hist: hist[instr] = []
    hist[instr] = [h for h in hist[instr] if h.get("fecha") != fecha]
    hist[instr].append({"fecha": fecha, "hora": hora, "snap": snap_rows})
    hist[instr] = sorted(hist[instr], key=lambda h: h.get("fecha", ""))[-365:]


# ═══════════════════════════════════════════════════════════════════════════════
# 1. LECAPS
# ═══════════════════════════════════════════════════════════════════════════════
LECAPS_COLS = [
    {"key":"ticker",  "label":"Especie",     "w": 8},
    {"key":"vto",     "label":"Vto.",         "w":11, "fmt":"DD/MM/YYYY"},
    {"key":"dias",    "label":"Días",         "w": 7, "align":"right"},
    {"key":"precio",  "label":"Precio Dirty", "w":12, "align":"right", "fmt":"#,##0.000"},
    {"key":"var",     "label":"Var %",        "w": 9, "align":"right", "fmt":"0.00%"},
    {"key":"tem",     "label":"TEM (%)",      "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"tna",     "label":"TNA (%)",      "w":10, "align":"right", "fmt":"0.0%"},
    {"key":"tirea",   "label":"TIREA (%)",    "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"volumen", "label":"Volumen",      "w":18, "align":"right", "fmt":"#,##0"},
    {"key":"vf",      "label":"Valor Final",  "w":12, "align":"right", "fmt":"#,##0.000"},
]

def _parse_lecaps(r, liq):
    vto   = _d(r.get("vencimiento") or r.get("fechaVencimiento"))
    tirea = _f(r.get("tea_bruta") or r.get("tirea"))
    tem   = _f(r.get("tem_bruta") or r.get("tem"))
    tna   = _f(r.get("tna_simple_bruta") or r.get("tna"))
    rv    = r.get("pct_change") if r.get("pct_change") is not None else r.get("variacionDiaria")
    if rv is None:
        p = _f(r.get("precio")); p0 = _f(r.get("precio_anterior") or r.get("precioAnterior"))
        rv = ((p/p0)-1)*100 if (p and p0 and p0 != 0) else None
    dias = r.get("dias_al_vto")
    if dias is None and vto and liq: dias = (vto - liq).days
    return {"ticker": (r.get("ticker") or "").strip(), "vto": vto,
            "dias": int(dias) if dias is not None else None,
            "precio": _f(r.get("precio")), "var": _pct(rv), "tem": tem, "tna": tna,
            "tirea": tirea, "volumen": _f(r.get("v") or r.get("volumen")),
            "vf": _f(r.get("valor_final") or r.get("valorFinal") or r.get("vf"))}

def _es_valido_lecaps(r, hoy):
    return _filtro_comun(_f(r.get("precio")), _f(r.get("v") or r.get("volumen")),
                         _f(r.get("tea_bruta") or r.get("tirea")),
                         _d(r.get("vencimiento") or r.get("fechaVencimiento")), hoy)

def build_lecaps(wb, hist, fecha_hoy, hora_str, fecha_dt, hora_dt):
    raw = api_get("/api/lecaps"); hoy = _d(fecha_hoy)
    liq = hoy + timedelta(days=1) if hoy else None
    rows = [_parse_lecaps(r, liq) for r in raw if _es_valido_lecaps(r, hoy)]
    rows.sort(key=lambda r: r.get("dias") or 9999)
    print(f"  {len(rows)} válidos  ({len(raw)-len(rows)} excluidos)")
    if not rows: return
    snap = [{"ticker":r["ticker"],"dias":r["dias"],"precio":r["precio"],"var":r["var"],"volumen":r["volumen"],"tem":r["tem"],"tna":r["tna"],"tirea":r["tirea"]} for r in rows]
    add_snapshot_to_hist(hist, "LECAPS", fecha_hoy, hora_str, snap)
    ws = wb.create_sheet("LECAPS"); ws.sheet_properties.tabColor = TAB_LECAPS
    write_meta(ws, "LECAPS / BONCAPS — Tasa Fija ARS", fecha_dt, hora_dt, liq)
    write_hdr(ws, 5, LECAPS_COLS)
    for i, r in enumerate(rows): write_row(ws, 6+i, LECAPS_COLS, r, alt=(i%2==1))
    ws.freeze_panes = "A6"

# ═══════════════════════════════════════════════════════════════════════════════
# 2. BONCER
# ═══════════════════════════════════════════════════════════════════════════════
BONCER_COLS = [
    {"key":"ticker",  "label":"Ticker",       "w": 9},
    {"key":"vto",     "label":"Vencimiento",  "w":13, "fmt":"DD/MM/YYYY"},
    {"key":"precio",  "label":"Precio",       "w":12, "align":"right", "fmt":"#,##0.00"},
    {"key":"var",     "label":"% Var Diaria", "w":11, "align":"right", "fmt":"0.00%"},
    {"key":"tirea",   "label":"TIREA (%)",    "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"tna",     "label":"TNA (%)",      "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"md",      "label":"MD",           "w": 8, "align":"right", "fmt":"0.00"},
    {"key":"vt",      "label":"VT",           "w":12, "align":"right", "fmt":"#,##0.00"},
    {"key":"paridad", "label":"Paridad",      "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"ai",      "label":"Int. Corrido", "w":12, "align":"right", "fmt":"#,##0.00"},
    {"key":"volumen", "label":"Volumen",      "w":18, "align":"right", "fmt":"#,##0"},
]

def _parse_boncer(r):
    return {"ticker": (r.get("ticker") or "").strip(),
            "vto":    _d(r.get("fecha") or r.get("fechaVencimiento") or r.get("vencimiento")),
            "precio": _f(r.get("precio_sucio") or r.get("precioSucio")),
            "var":    _pct(r.get("variacion_diaria") or r.get("variacionDiaria")),
            "tirea":  _f(r.get("tirea")), "tna": _f(r.get("tna")),
            "md":     _f(r.get("modified_duration") or r.get("modifiedDuration")),
            "vt":     _f(r.get("valorTecnico")), "paridad": _f(r.get("paridad")),
            "ai":     _f(r.get("interesesCorridos")), "volumen": _f(r.get("volumen"))}

def _es_valido_boncer(r, hoy):
    return _filtro_comun(_f(r.get("precio_sucio") or r.get("precioSucio")), _f(r.get("volumen")),
                         _f(r.get("tirea")),
                         _d(r.get("fecha") or r.get("fechaVencimiento") or r.get("vencimiento")), hoy)

def build_boncer(wb, hist, fecha_hoy, hora_str, fecha_dt, hora_dt):
    data = api_get_obj("/api/boncer")
    raw  = data.get("rows",[]) if isinstance(data,dict) else (data if isinstance(data,list) else [])
    hoy  = _d(fecha_hoy); liq = hoy + timedelta(days=1) if hoy else None
    rows = [_parse_boncer(r) for r in raw if _es_valido_boncer(r, hoy)]
    rows.sort(key=lambda r: r.get("vto") or date.max)
    print(f"  {len(rows)} válidos  ({len(raw)-len(rows)} excluidos)")
    if not rows: return
    snap = [{"ticker":r["ticker"],"md":r["md"],"precio":r["precio"],"var":r["var"],"volumen":r["volumen"],"tirea":r["tirea"],"tna":r["tna"]} for r in rows]
    add_snapshot_to_hist(hist, "BONCER", fecha_hoy, hora_str, snap)
    ws = wb.create_sheet("BONCER"); ws.sheet_properties.tabColor = TAB_BONCER
    write_meta(ws, "BONCER — Bonos CER-Linked ARS", fecha_dt, hora_dt, liq)
    write_hdr(ws, 5, BONCER_COLS)
    for i, r in enumerate(rows): write_row(ws, 6+i, BONCER_COLS, r, alt=(i%2==1))
    ws.freeze_panes = "A6"

# ═══════════════════════════════════════════════════════════════════════════════
# 3. DLK
# ═══════════════════════════════════════════════════════════════════════════════
DLK_COLS = [
    {"key":"ticker",     "label":"Ticker",      "w": 9},
    {"key":"emisor",     "label":"Emisor",       "w":20},
    {"key":"vto",        "label":"Vencimiento",  "w":13, "fmt":"DD/MM/YYYY"},
    {"key":"precio_ars", "label":"Precio ARS",   "w":14, "align":"right", "fmt":"#,##0.00"},
    {"key":"precio_usd", "label":"Precio USD",   "w":12, "align":"right", "fmt":"#,##0.0000"},
    {"key":"var",        "label":"% Var Diaria", "w":11, "align":"right", "fmt":"0.00%"},
    {"key":"tirea",      "label":"TIREA (%)",    "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"tna",        "label":"TNA (%)",      "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"md",         "label":"MD",           "w": 8, "align":"right", "fmt":"0.00"},
    {"key":"vt",         "label":"VT",           "w": 8, "align":"right", "fmt":"#,##0.00"},
    {"key":"paridad",    "label":"Paridad",      "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"ai",         "label":"Int. Corrido", "w":12, "align":"right", "fmt":"#,##0.0000"},
    {"key":"volumen",    "label":"Volumen",      "w":15, "align":"right", "fmt":"#,##0"},
    {"key":"a3500",      "label":"A3500",        "w":10, "align":"right", "fmt":"#,##0.00"},
]

def _parse_dlk(r):
    return {"ticker": (r.get("ticker") or "").strip(), "emisor": r.get("emisor") or "",
            "vto": _d(r.get("fecha") or r.get("fechaVencimiento")),
            "precio_ars": _f(r.get("precio_ars")),
            "precio_usd": _f(r.get("precio_limpio") or r.get("precioUsd")),
            "var": _pct(r.get("variacion_diaria")),
            "tirea": _f(r.get("tirea")), "tna": _f(r.get("tna")),
            "md": _f(r.get("modified_duration")), "vt": _f(r.get("vt")),
            "paridad": _f(r.get("paridad")), "ai": _f(r.get("ai_usd")),
            "volumen": _f(r.get("volumen")), "a3500": _f(r.get("a3500"))}

def _es_valido_dlk(r, hoy):
    return _filtro_comun(_f(r.get("precio_ars")), _f(r.get("volumen")), _f(r.get("tirea")),
                         _d(r.get("fecha") or r.get("fechaVencimiento")), hoy)

def build_dlk(wb, hist, fecha_hoy, hora_str, fecha_dt, hora_dt):
    data = api_get_obj("/api/dlk")
    raw  = data.get("rows",[]) if isinstance(data,dict) else (data if isinstance(data,list) else [])
    hoy  = _d(fecha_hoy); liq = hoy + timedelta(days=1) if hoy else None
    rows = [_parse_dlk(r) for r in raw if _es_valido_dlk(r, hoy)]
    rows.sort(key=lambda r: r.get("vto") or date.max)
    print(f"  {len(rows)} válidos  ({len(raw)-len(rows)} excluidos)")
    if not rows: return
    snap = [{"ticker":r["ticker"],"md":r["md"],"precio_ars":r["precio_ars"],"var":r["var"],"volumen":r["volumen"],"tirea":r["tirea"],"tna":r["tna"]} for r in rows]
    add_snapshot_to_hist(hist, "DLK", fecha_hoy, hora_str, snap)
    ws = wb.create_sheet("DLK"); ws.sheet_properties.tabColor = TAB_DLK
    write_meta(ws, "Dólar Linked — Bonos atados al A3500", fecha_dt, hora_dt, liq)
    write_hdr(ws, 5, DLK_COLS)
    for i, r in enumerate(rows): write_row(ws, 6+i, DLK_COLS, r, alt=(i%2==1))
    ws.freeze_panes = "A6"

# ═══════════════════════════════════════════════════════════════════════════════
# 4. TAMAR
# ═══════════════════════════════════════════════════════════════════════════════
TAMAR_COLS = [
    {"key":"ticker",      "label":"Ticker",            "w": 9},
    {"key":"vto",         "label":"Vencimiento",       "w":13, "fmt":"DD/MM/YYYY"},
    {"key":"dias_vto",    "label":"Días al Vto",       "w": 9, "align":"right"},
    {"key":"precio",      "label":"Precio",            "w":10, "align":"right", "fmt":"#,##0.00"},
    {"key":"var",         "label":"% Var Diaria",      "w":11, "align":"right", "fmt":"0.00%"},
    {"key":"tem_tamar",   "label":"TEM TAMAR",         "w":11, "align":"right", "fmt":"0.00%"},
    {"key":"tna_tamar",   "label":"TNA TAMAR",         "w":11, "align":"right", "fmt":"0.00%"},
    {"key":"tirea_tamar", "label":"TIREA TAMAR",       "w":11, "align":"right", "fmt":"0.00%"},
    {"key":"vf_tamar",    "label":"Valor Final",       "w":11, "align":"right", "fmt":"#,##0.00"},
    {"key":"tamar_aplic", "label":"TAMAR Aplic (TNA)", "w":15, "align":"right", "fmt":"0.00%"},
    {"key":"spread_bps",  "label":"Spread (bps)",      "w":12, "align":"right", "fmt":"#,##0.00"},
    {"key":"volumen",     "label":"Volumen",           "w":18, "align":"right", "fmt":"#,##0"},
]

def _parse_tamar(r):
    return {"ticker": (r.get("ticker") or "").strip(),
            "vto": _d(r.get("fechaVencimiento")), "dias_vto": r.get("diasAlVto"),
            "precio": _f(r.get("precioSucio")), "var": _pct(r.get("variacionDiaria")),
            "tem_tamar": _f(r.get("temTamar")), "tna_tamar": _f(r.get("tnaTamar")),
            "tirea_tamar": _f(r.get("tireaTamar")), "vf_tamar": _f(r.get("valorFinalTamar")),
            "tamar_aplic": _f(r.get("tamarTnaAplicable")), "spread_bps": _f(r.get("spreadTamarBps")),
            "volumen": _f(r.get("volumen"))}

def _es_valido_tamar(r, hoy):
    return _filtro_comun(_f(r.get("precioSucio")), _f(r.get("volumen")), _f(r.get("tireaTamar")),
                         _d(r.get("fechaVencimiento")), hoy)

def build_tamar(wb, hist, fecha_hoy, hora_str, fecha_dt, hora_dt):
    raw = api_get("/api/tamar"); hoy = _d(fecha_hoy); liq = hoy + timedelta(days=1) if hoy else None
    rows = [_parse_tamar(r) for r in raw if _es_valido_tamar(r, hoy)]
    rows.sort(key=lambda r: r.get("vto") or date.max)
    print(f"  {len(rows)} válidos  ({len(raw)-len(rows)} excluidos)")
    if not rows: return
    snap = [{"ticker":r["ticker"],"dias_vto":r["dias_vto"],"precio":r["precio"],"var":r["var"],"volumen":r["volumen"],"tna_tamar":r["tna_tamar"],"tirea_tamar":r["tirea_tamar"]} for r in rows]
    add_snapshot_to_hist(hist, "TAMAR", fecha_hoy, hora_str, snap)
    ws = wb.create_sheet("TAMAR"); ws.sheet_properties.tabColor = TAB_TAMAR
    write_meta(ws, "TAMAR — Bonos tasa variable TAMAR", fecha_dt, hora_dt, liq)
    write_hdr(ws, 5, TAMAR_COLS)
    for i, r in enumerate(rows): write_row(ws, 6+i, TAMAR_COLS, r, alt=(i%2==1))
    ws.freeze_panes = "A6"

# ═══════════════════════════════════════════════════════════════════════════════
# 5. DUALES
# ═══════════════════════════════════════════════════════════════════════════════
DUAL_COLS = [
    {"key":"ticker",      "label":"Ticker",            "w": 9},
    {"key":"vto",         "label":"Vencimiento",       "w":13, "fmt":"DD/MM/YYYY"},
    {"key":"precio",      "label":"Precio",            "w":10, "align":"right", "fmt":"#,##0.00"},
    {"key":"var",         "label":"% Var Diaria",      "w":11, "align":"right", "fmt":"0.00%"},
    {"key":"tem_tamar",   "label":"TEM TAMAR",         "w":11, "align":"right", "fmt":"0.00%"},
    {"key":"tna_tamar",   "label":"TNA TAMAR",         "w":11, "align":"right", "fmt":"0.00%"},
    {"key":"tirea_tamar", "label":"TIREA TAMAR",       "w":11, "align":"right", "fmt":"0.00%"},
    {"key":"vf_tamar",    "label":"VF TAMAR",          "w":10, "align":"right", "fmt":"#,##0.00"},
    {"key":"dias_vto",    "label":"Días al Vto",       "w": 9, "align":"right"},
    {"key":"tem_fija",    "label":"TEM Fija",          "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"tna_fija",    "label":"TNA Fija",          "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"tirea_fija",  "label":"TIREA Fija",        "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"vf_fija",     "label":"VF Fija",           "w":10, "align":"right", "fmt":"#,##0.00"},
    {"key":"tamar_aplic", "label":"TAMAR Aplic (TNA)", "w":15, "align":"right", "fmt":"0.00%"},
    {"key":"pata",        "label":"Pata Elegida",      "w":12},
    {"key":"spread_bps",  "label":"Spread (bps)",      "w":12, "align":"right", "fmt":"#,##0.00"},
    {"key":"volumen",     "label":"Volumen",           "w":18, "align":"right", "fmt":"#,##0"},
]

def _parse_dual(r):
    return {"ticker": (r.get("ticker") or "").strip(),
            "vto": _d(r.get("fechaVencimiento")), "dias_vto": r.get("diasAlVto"),
            "precio": _f(r.get("precioSucio")), "var": _pct(r.get("variacionDiaria")),
            "tem_tamar": _f(r.get("temTamar")), "tna_tamar": _f(r.get("tnaTamar")),
            "tirea_tamar": _f(r.get("tireaTamar")), "vf_tamar": _f(r.get("valorFinalTamar")),
            "tem_fija": _f(r.get("temFija")), "tna_fija": _f(r.get("tnaFija")),
            "tirea_fija": _f(r.get("tireaFija")), "vf_fija": _f(r.get("valorFinalFija")),
            "tamar_aplic": _f(r.get("tamarTnaAplicable")), "pata": r.get("pataElegida") or "",
            "spread_bps": _f(r.get("spreadTamarBps")), "volumen": _f(r.get("volumen"))}

def _es_valido_dual(r, hoy):
    return _filtro_comun(_f(r.get("precioSucio")), _f(r.get("volumen")),
                         _f(r.get("tireaTamar") or r.get("tireaFija")),
                         _d(r.get("fechaVencimiento")), hoy)

def build_duales(wb, hist, fecha_hoy, hora_str, fecha_dt, hora_dt):
    raw = api_get("/api/duales"); hoy = _d(fecha_hoy); liq = hoy + timedelta(days=1) if hoy else None
    rows = [_parse_dual(r) for r in raw if _es_valido_dual(r, hoy)]
    rows.sort(key=lambda r: r.get("vto") or date.max)
    print(f"  {len(rows)} válidos  ({len(raw)-len(rows)} excluidos)")
    if not rows: return
    snap = [{"ticker":r["ticker"],"dias_vto":r["dias_vto"],"precio":r["precio"],"var":r["var"],"volumen":r["volumen"],"tirea_tamar":r["tirea_tamar"],"tirea_fija":r["tirea_fija"]} for r in rows]
    add_snapshot_to_hist(hist, "DUAL", fecha_hoy, hora_str, snap)
    ws = wb.create_sheet("DUALES"); ws.sheet_properties.tabColor = TAB_DUAL
    write_meta(ws, "Duales — TAMAR vs Tasa Fija", fecha_dt, hora_dt, liq)
    write_hdr(ws, 5, DUAL_COLS)
    for i, r in enumerate(rows): write_row(ws, 6+i, DUAL_COLS, r, alt=(i%2==1))
    ws.freeze_panes = "A6"

# ═══════════════════════════════════════════════════════════════════════════════
# 6-7. ONs
# ═══════════════════════════════════════════════════════════════════════════════
ONS_COLS = [
    {"key":"ticker",      "label":"Ticker",        "w": 9},
    {"key":"emisor",      "label":"Empresa",        "w":22},
    {"key":"vto",         "label":"Vencimiento",    "w":13, "fmt":"DD/MM/YYYY"},
    {"key":"precio_ars",  "label":"Precio ARS",     "w":13, "align":"right", "fmt":"#,##0"},
    {"key":"var",         "label":"% Var Diaria",   "w":13, "align":"right", "fmt":"0.00%"},
    {"key":"precio_d",    "label":"Precio D",       "w":13, "align":"right", "fmt":"#,##0.00"},
    {"key":"var_d",       "label":"% Var D",        "w":13, "align":"right", "fmt":"0.00%"},
    {"key":"tirea_d",     "label":"TIREA D (%)",    "w":14, "align":"right", "fmt":"0.00%"},
    {"key":"tna_d",       "label":"TNA D (%)",      "w":13, "align":"right", "fmt":"0.00%"},
    {"key":"md_d",        "label":"MD D",           "w":10, "align":"right", "fmt":"#,##0.00"},
    {"key":"precio_c",    "label":"Precio C",       "w":13, "align":"right", "fmt":"#,##0.00"},
    {"key":"var_c",       "label":"% Var C",        "w":13, "align":"right", "fmt":"0.00%"},
    {"key":"tirea_c",     "label":"TIREA C (%)",    "w":14, "align":"right", "fmt":"0.00%"},
    {"key":"tna_c",       "label":"TNA C (%)",      "w":13, "align":"right", "fmt":"0.00%"},
    {"key":"md_c",        "label":"MD C",           "w":10, "align":"right", "fmt":"#,##0.00"},
    {"key":"vt",          "label":"VT",             "w": 9, "align":"right", "fmt":"#,##0.00"},
    {"key":"paridad_d",   "label":"Paridad D",      "w":13, "align":"right", "fmt":"0.00%"},
    {"key":"ai",          "label":"Int. Corrido",   "w":13, "align":"right", "fmt":"#,##0.00"},
    {"key":"calificacion","label":"Calificación",   "w":14},
    {"key":"sector",      "label":"Sector",         "w":16},
    {"key":"volumen",     "label":"Volumen",        "w":18, "align":"right", "fmt":"#,##0"},
]

def _parse_ons(r, liq):
    p = _f(r.get("paridad") or r.get("paridadMep"))
    return {"ticker": (r.get("ticker") or "").strip(),
            "emisor": r.get("emisor") or r.get("empresa") or "",
            "vto": _d(r.get("fechaVencimiento") or r.get("vencimiento")),
            "precio_ars": _f(r.get("precioSucioArs") or r.get("precioArs")),
            "var":   _pct(r.get("variacionDiaria") or r.get("variacion_diaria")),
            "precio_d": _f(r.get("precioD") or r.get("precioMep")),
            "var_d":    _pct(r.get("variacionDiariaD") or r.get("variacionDiaria")),
            "tirea_d":  _f(r.get("tireaMep")),
            "tna_d":    _f(r.get("tnaMep")),
            "md_d":     _f(r.get("mdMep")),
            "precio_c": _f(r.get("precioC") or r.get("precioCcl")),
            "var_c":    _pct(r.get("variacionDiariaC") or r.get("variacionDiaria")),
            "tirea_c":  _f(r.get("tireaCcl")),
            "tna_c":    _f(r.get("tnaCcl")),
            "md_c":     _f(r.get("mdCcl")),
            "vt":       _f(r.get("vt") or r.get("valorTecnico")),
            "paridad_d": p if p and p <= 3 else (p/100 if p else None),
            "ai":       _f(r.get("ai") or r.get("interesesCorridos")),
            "calificacion": r.get("calificacionRiesgo") or r.get("calificacion") or "",
            "sector":   r.get("sector") or "",
            "volumen":  _f(r.get("volumen") or r.get("v"))}

def _es_valido_ons(r, hoy):
    return _filtro_comun(_f(r.get("precioD") or r.get("precioMep") or r.get("precioMEP")),
                         _f(r.get("volumen") or r.get("v")),
                         _f(r.get("tireaMep") or r.get("tireaMEP")),
                         _d(r.get("fechaVencimiento") or r.get("vencimiento")), hoy)

def build_ons(wb, hist, endpoint, instr_key, sheet_name, titulo, tab_color,
              fecha_hoy, hora_str, fecha_dt, hora_dt):
    raw = api_get(endpoint); hoy = _d(fecha_hoy); liq = hoy + timedelta(days=1) if hoy else None
    rows = [_parse_ons(r, liq) for r in raw if _es_valido_ons(r, hoy)]
    rows.sort(key=lambda r: r.get("vto") or date.max)
    print(f"  {len(rows)} válidos  ({len(raw)-len(rows)} excluidos)")
    if not rows: return
    snap = [{"ticker":r["ticker"],"md":r["md_d"],"precio":r["precio_d"],"var":r["var"],"volumen":r["volumen"],"tirea":r["tirea_d"],"tirea_ccl":r["tirea_c"]} for r in rows]
    add_snapshot_to_hist(hist, instr_key, fecha_hoy, hora_str, snap)
    ws = wb.create_sheet(sheet_name); ws.sheet_properties.tabColor = tab_color
    write_meta(ws, titulo, fecha_dt, hora_dt, liq)
    write_hdr(ws, 5, ONS_COLS)
    for i, r in enumerate(rows): write_row(ws, 6+i, ONS_COLS, r, alt=(i%2==1))
    ws.freeze_panes = "A6"

# ═══════════════════════════════════════════════════════════════════════════════
# 8. HD SOBERANOS
# ═══════════════════════════════════════════════════════════════════════════════
HDSOB_COLS = [
    {"key":"ticker",    "label":"Ticker",      "w": 9},
    {"key":"categoria", "label":"Categoría",   "w":12},
    {"key":"vto",       "label":"Vencimiento", "w":13, "fmt":"DD/MM/YYYY"},
    {"key":"precio",  "label":"Precio",       "w":10, "align":"right", "fmt":"#,##0.00"},
    {"key":"var",     "label":"% Var Diaria", "w":11, "align":"right", "fmt":"0.00%"},
    {"key":"tirea",   "label":"TIREA (%)",    "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"tna",     "label":"TNA (%)",      "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"md",      "label":"MD",           "w": 8, "align":"right", "fmt":"0.00"},
    {"key":"vt",      "label":"VT",           "w":10, "align":"right", "fmt":"#,##0.00"},
    {"key":"paridad", "label":"Paridad",      "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"ai",      "label":"Int. Corrido", "w":12, "align":"right", "fmt":"#,##0.000"},
    {"key":"volumen", "label":"Volumen",      "w":15, "align":"right", "fmt":"#,##0"},
]

def _categoria_hdsob(r):
    emisor = (r.get("emisor") or "").strip().lower()
    ley    = (r.get("ley")    or "").strip().lower()
    if "bcra" in emisor: return "Bopreales"
    if "gobierno nacional" in emisor or "gobierno" in emisor:
        if "nueva york" in ley or "new york" in ley or "york" in ley: return "Globales"
        if "nacional" in ley or "argentina" in ley: return "Bonares"
        t = (r.get("ticker") or "").upper()
        if t.startswith("GD") or t.startswith("GE") or t.startswith("AE"): return "Globales"
        if t.startswith("AL") or t.startswith("AO") or t.startswith("AN") or t.startswith("AC"): return "Bonares"
    t = (r.get("ticker") or "").upper()
    if t.startswith("BP"): return "Bopreales"
    if t.startswith("GD") or t.startswith("GE") or t.startswith("AE"): return "Globales"
    if t.startswith("AL") or t.startswith("AO") or t.startswith("AN") or t.startswith("AC"): return "Bonares"
    return "Otros"

_CAT_ORDER = {"Globales": 0, "Bonares": 1, "Bopreales": 2, "Otros": 3}

def _parse_hdsob(r):
    return {"ticker": (r.get("ticker") or "").strip(),
            "categoria": _categoria_hdsob(r),
            "vto": _d(r.get("fechaVencimiento")), "precio": _f(r.get("precioSucio")),
            "var": _pct(r.get("variacionDiaria")), "tirea": _f(r.get("tirea")),
            "tna": _f(r.get("tna")), "md": _f(r.get("duracionMod")),
            "vt": _f(r.get("valorTecnico")), "paridad": _f(r.get("paridad")),
            "ai": _f(r.get("ai")), "volumen": _f(r.get("volumen"))}

def _es_valido_hdsob(r, hoy):
    return _filtro_comun(_f(r.get("precioSucio")), _f(r.get("volumen")), _f(r.get("tirea")),
                         _d(r.get("fechaVencimiento")), hoy)

def build_hdsob(wb, hist, fecha_hoy, hora_str, fecha_dt, hora_dt):
    raw = api_get("/api/bonoshd"); hoy = _d(fecha_hoy); liq = hoy + timedelta(days=1) if hoy else None
    rows = [_parse_hdsob(r) for r in raw if _es_valido_hdsob(r, hoy)]
    rows.sort(key=lambda r: (_CAT_ORDER.get(r.get("categoria","Otros"), 3), r.get("vto") or date.max))
    print(f"  {len(rows)} válidos  ({len(raw)-len(rows)} excluidos)")
    if not rows: return
    snap = [{"ticker":r["ticker"],"md":r["md"],"precio":r["precio"],"var":r["var"],"volumen":r["volumen"],"tirea":r["tirea"],"tna":r["tna"]} for r in rows]
    add_snapshot_to_hist(hist, "HDSOB", fecha_hoy, hora_str, snap)
    ws = wb.create_sheet("HD_Sob"); ws.sheet_properties.tabColor = TAB_HDSOB
    write_meta(ws, "Bonos Hard Dollar Soberanos (Bonares + Globales)", fecha_dt, hora_dt, liq)
    write_hdr(ws, 5, HDSOB_COLS)
    for i, r in enumerate(rows): write_row(ws, 6+i, HDSOB_COLS, r, alt=(i%2==1))
    ws.freeze_panes = "A6"

# ═══════════════════════════════════════════════════════════════════════════════
# 9. BOTE
# ═══════════════════════════════════════════════════════════════════════════════
BOTE_COLS = [
    {"key":"ticker",        "label":"Ticker",        "w": 9},
    {"key":"vto",           "label":"Vencimiento",   "w":13, "fmt":"DD/MM/YYYY"},
    {"key":"precio_sucio",  "label":"Precio Sucio",  "w":12, "align":"right", "fmt":"#,##0.00"},
    {"key":"precio_limpio", "label":"Precio Limpio", "w":12, "align":"right", "fmt":"#,##0.00"},
    {"key":"var",           "label":"% Var Diaria",  "w":11, "align":"right", "fmt":"0.00%"},
    {"key":"tirea",         "label":"TIREA (%)",     "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"tna",           "label":"TNA (%)",       "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"md",            "label":"MD",            "w": 8, "align":"right", "fmt":"0.00"},
    {"key":"wal",           "label":"WAL",           "w": 8, "align":"right", "fmt":"0.00"},
    {"key":"vt",            "label":"VT",            "w":10, "align":"right", "fmt":"#,##0.00"},
    {"key":"vr",            "label":"VR",            "w": 8, "align":"right", "fmt":"#,##0.00"},
    {"key":"paridad",       "label":"Paridad",       "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"ai",            "label":"Int. Corrido",  "w":12, "align":"right", "fmt":"#,##0.000"},
    {"key":"cupon",         "label":"Cupón %",       "w":10, "align":"right", "fmt":"0.00%"},
    {"key":"volumen",       "label":"Volumen",       "w":15, "align":"right", "fmt":"#,##0"},
]

def _parse_bote(r):
    return {"ticker": (r.get("ticker") or "").strip(), "emisor": r.get("emisor") or "",
            "vto": _d(r.get("fechaVencimiento")),
            "precio_sucio": _f(r.get("precioSucio")), "precio_limpio": _f(r.get("precioLimpio")),
            "var": _pct(r.get("variacionDiaria")), "tirea": _f(r.get("tirea")),
            "tna": _f(r.get("tna")), "md": _f(r.get("duracionMod")), "wal": _f(r.get("wal")),
            "vt": _f(r.get("valorTecnico")), "vr": _f(r.get("vr")),
            "paridad": _f(r.get("paridad")), "ai": _f(r.get("ai")),
            "cupon": _f(r.get("cuponFijoPct")), "volumen": _f(r.get("volumen"))}

def _es_valido_bote(r, hoy):
    return _filtro_comun(_f(r.get("precioSucio")), _f(r.get("volumen")), _f(r.get("tirea")),
                         _d(r.get("fechaVencimiento")), hoy)

def build_bote(wb, hist, fecha_hoy, hora_str, fecha_dt, hora_dt):
    raw = api_get("/api/bote"); hoy = _d(fecha_hoy); liq = hoy + timedelta(days=1) if hoy else None
    rows = [_parse_bote(r) for r in raw if _es_valido_bote(r, hoy)]
    rows.sort(key=lambda r: r.get("vto") or date.max)
    print(f"  {len(rows)} válidos  ({len(raw)-len(rows)} excluidos)")
    if not rows: return
    snap = [{"ticker":r["ticker"],"md":r["md"],"precio_sucio":r["precio_sucio"],"var":r["var"],"volumen":r["volumen"],"tirea":r["tirea"],"tna":r["tna"],"wal":r["wal"]} for r in rows]
    add_snapshot_to_hist(hist, "BOTE", fecha_hoy, hora_str, snap)
    ws = wb.create_sheet("BOTE"); ws.sheet_properties.tabColor = TAB_BOTE
    write_meta(ws, "BOTE — Bonos Tasa Fija USD", fecha_dt, hora_dt, liq)
    write_hdr(ws, 5, BOTE_COLS)
    for i, r in enumerate(rows): write_row(ws, 6+i, BOTE_COLS, r, alt=(i%2==1))
    ws.freeze_panes = "A6"

# ═══════════════════════════════════════════════════════════════════════════════
# HISTORIA  v3
# ═══════════════════════════════════════════════════════════════════════════════
def build_historia(wb, hist):
    ws = wb.create_sheet("HISTORIA")
    ws.sheet_properties.tabColor = TAB_HIST
    ws.sheet_view.showGridLines   = False
    for r in range(1, 3000):
        for c in range(1, 14):
            ws.cell(r, c).fill = _fill(BG_TITLE)

    ws["A1"].value = "Historia de cierres diarios  v3.0"
    ws["A1"].font  = Font(name="Arial", bold=True, color=FG_TITLE, size=12)
    ws["A1"].fill  = _fill(BG_TITLE)
    ws["A2"].value = "Extra: DUAL=TIREA Fija  |  ONs=TIREA CCL  |  BOTE=WAL"
    ws["A2"].font  = _font(FG_MUTED, size=7); ws["A2"].fill = _fill(BG_TITLE)

    hdrs   = ["Instrumento","Fecha","Hora","Ticker"] + HIST_FIXED_HDRS
    widths = [14, 11, 7, 10] + HIST_FIXED_WIDTHS
    for i, (h, w) in enumerate(zip(hdrs, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(3, i)
        c.value = h; c.fill = _fill(BG_HDR); c.font = _font(FG_WHITE, bold=True)
        c.alignment = _center(); c.border = _border()

    all_entries = []
    for instr, entries in hist.items():
        m = HIST_MAP.get(instr)
        if not m: continue
        for entry in entries:
            all_entries.append((instr, entry))
    all_entries.sort(key=lambda x: (x[1].get("fecha",""), x[0]), reverse=True)

    row_idx = 4
    for instr, entry in all_entries:
        m = HIST_MAP.get(instr)
        if not m: continue
        fecha     = entry.get("fecha","")
        hora      = entry.get("hora","")
        snap_rows = entry.get("snap") or []
        raw_rows  = entry.get("_raw") or []
        use_old   = not snap_rows and bool(raw_rows)
        rows_src  = snap_rows if not use_old else raw_rows
        for r in rows_src:
            if use_old:
                ticker  = r.get("_ticker","")
                hora_r  = str(r.get("_hora") or hora or "")
                vals_raw = r.get("_vals") or []
                v_dim   = vals_raw[0] if len(vals_raw) > 0 else None
                v_precio = None
                v_var    = None
                v_vol    = None
                v_tem   = vals_raw[1] if len(vals_raw) > 1 else None
                v_tna   = vals_raw[2] if len(vals_raw) > 2 else None
                v_tirea = vals_raw[3] if len(vals_raw) > 3 else None
                v_extra = vals_raw[4] if len(vals_raw) > 4 else None
            else:
                ticker   = r.get("ticker","")
                hora_r   = hora
                v_dim    = r.get(m["dim"])
                v_precio = r.get(m["precio"]) if m.get("precio") else None
                v_var    = r.get(m["var"])    if m.get("var")    else None
                v_vol    = r.get(m["vol"])    if m.get("vol")    else None
                v_tem    = r.get(m["tem"])    if m["tem"]        else None
                v_tna    = r.get(m["tna"])    if m["tna"]        else None
                v_tirea  = r.get(m["tirea"])  if m["tirea"]      else None
                v_extra  = r.get(m["extra"])  if m["extra"]      else None

            ws.cell(row_idx,1).value = instr
            ws.cell(row_idx,2).value = fecha
            ws.cell(row_idx,3).value = hora_r
            ws.cell(row_idx,4).value = ticker

            c5 = ws.cell(row_idx,5); c5.value = v_precio
            if v_precio is not None: c5.number_format = m["precio_fmt"]
            c6 = ws.cell(row_idx,6); c6.value = v_var
            if v_var is not None: c6.number_format = "0.00%"
            c7 = ws.cell(row_idx,7); c7.value = v_tem
            if v_tem is not None: c7.number_format = "0.00%"
            c8 = ws.cell(row_idx,8); c8.value = v_tna
            if v_tna is not None: c8.number_format = "0.00%"
            c9 = ws.cell(row_idx,9); c9.value = v_tirea
            if v_tirea is not None: c9.number_format = "0.00%"
            c10 = ws.cell(row_idx,10); c10.value = v_dim
            if v_dim is not None: c10.number_format = m["dim_fmt"]
            c11 = ws.cell(row_idx,11); c11.value = v_vol
            if v_vol is not None: c11.number_format = "#,##0"
            c12 = ws.cell(row_idx,12); c12.value = v_extra
            if v_extra is not None and m["extra_fmt"]:
                c12.number_format = m["extra_fmt"]

            alt = (row_idx % 2 == 0)
            for ci in range(1, 13):
                c = ws.cell(row_idx, ci)
                c.fill = _fill(BG_ALT if alt else BG_ROW)
                c.font = _font(FG_DATA, size=8)
                c.border = _border(); c.alignment = _center()
            row_idx += 1

    ws.freeze_panes = "A4"
    print(f"  {row_idx-4} registros en historia")


# ═══════════════════════════════════════════════════════════════════════════════
# DÓLAR
# ═══════════════════════════════════════════════════════════════════════════════
DOLAR_URL = "https://api.alquimiaconsultora.com/output/dolares.json"

DOLAR_MAIN_COLS = [
    {"key": "tipo",      "label": "Tipo",       "w": 16},
    {"key": "precio",    "label": "Precio",     "w": 12, "align": "right", "fmt": "#,##0.00"},
    {"key": "compra",    "label": "Compra",     "w": 12, "align": "right", "fmt": "#,##0.00"},
    {"key": "venta",     "label": "Venta",      "w": 12, "align": "right", "fmt": "#,##0.00"},
    {"key": "variacion", "label": "Var %",      "w":  9, "align": "right", "fmt": "0.00%"},
]
DOLAR_MEP_CCL_COLS = [
    {"key": "tipo",      "label": "Tipo",       "w": 16},
    {"key": "bono",      "label": "Bono",       "w": 10},
    {"key": "plazo",     "label": "Plazo",      "w":  8},
    {"key": "precio",    "label": "Precio",     "w": 12, "align": "right", "fmt": "#,##0.00"},
    {"key": "variacion", "label": "Var %",      "w":  9, "align": "right", "fmt": "0.00%"},
]
DOLAR_CRIPTO_COLS = [
    {"key": "tipo",      "label": "Cripto",     "w": 10},
    {"key": "compra",    "label": "Compra",     "w": 12, "align": "right", "fmt": "#,##0.00"},
    {"key": "venta",     "label": "Venta",      "w": 12, "align": "right", "fmt": "#,##0.00"},
    {"key": "variacion", "label": "Var %",      "w":  9, "align": "right", "fmt": "0.00%"},
]

def _pct_direct(v):
    x = _f(v)
    return None if x is None else x / 100.0

def build_dolar(wb, fecha_dt, hora_dt):
    print(f"  Consultando {DOLAR_URL}...")
    try:
        r = requests.get(DOLAR_URL, timeout=30); r.raise_for_status()
        payload = r.json()
    except Exception as e:
        print(f"  [ERROR] No se pudo obtener datos de dólar: {e}"); return

    data = payload.get("data", {})
    if not data:
        print("  [WARN] Respuesta vacía, se omite solapa DÓLAR."); return

    try:
        ws = wb.create_sheet("DÓLAR")
    except Exception:
        ws = wb.create_sheet("DOLAR")
    ws.sheet_properties.tabColor = TAB_DOLAR

    try:
        write_meta(ws, "Dólar — Cotizaciones del Día", fecha_dt, hora_dt, None)

        TIPOS_SIMPLES = [
            ("Mayorista", "mayorista"), ("Oficial", "oficial"),
            ("Ahorro", "ahorro"), ("Tarjeta", "tarjeta"), ("Blue", "blue"),
        ]
        r5 = ws.cell(5, 1); r5.value = "Cotizaciones Generales"
        r5.font = _font(FG_TITLE, bold=True, size=9); r5.fill = _fill(BG_TITLE)
        ws.row_dimensions[5].height = 14

        write_hdr(ws, 6, DOLAR_MAIN_COLS)
        row_idx = 7
        for nombre, clave in TIPOS_SIMPLES:
            d = data.get(clave)
            if not d: continue
            fila = {"tipo": nombre, "precio": _f(d.get("price")),
                    "compra": _f(d.get("bid")), "venta": _f(d.get("ask")),
                    "variacion": _pct_direct(d.get("variation"))}
            write_row(ws, row_idx, DOLAR_MAIN_COLS, fila, alt=(row_idx % 2 == 0))
            row_idx += 1

        row_idx += 1
        r_sub = ws.cell(row_idx, 1); r_sub.value = "Dólar Cripto"
        r_sub.font = _font(FG_TITLE, bold=True, size=9); r_sub.fill = _fill(BG_TITLE)
        ws.row_dimensions[row_idx].height = 14; row_idx += 1

        write_hdr(ws, row_idx, DOLAR_CRIPTO_COLS); row_idx += 1
        cripto = data.get("cripto", {})
        for clave in ["ccb", "usdt", "usdc"]:
            d = cripto.get(clave)
            if not d: continue
            fila = {"tipo": clave.upper(), "compra": _f(d.get("bid")),
                    "venta": _f(d.get("ask")), "variacion": _pct_direct(d.get("variation"))}
            write_row(ws, row_idx, DOLAR_CRIPTO_COLS, fila, alt=(row_idx % 2 == 0))
            row_idx += 1

        row_idx += 1
        r_sub2 = ws.cell(row_idx, 1); r_sub2.value = "MEP / CCL"
        r_sub2.font = _font(FG_TITLE, bold=True, size=9); r_sub2.fill = _fill(BG_TITLE)
        ws.row_dimensions[row_idx].height = 14; row_idx += 1

        write_hdr(ws, row_idx, DOLAR_MEP_CCL_COLS); row_idx += 1
        for tipo_key, tipo_lbl in [("mep", "MEP"), ("ccl", "CCL")]:
            grupo = data.get(tipo_key, {})
            for bono_key in ["al30", "gd30", "letras", "bpo27"]:
                d_bono = grupo.get(bono_key)
                if not d_bono: continue
                for plazo_key in ["24hs", "ci"]:
                    d_plazo = d_bono.get(plazo_key)
                    if not d_plazo: continue
                    fila = {"tipo": tipo_lbl, "bono": bono_key.upper(), "plazo": plazo_key,
                            "precio": _f(d_plazo.get("price")),
                            "variacion": _pct_direct(d_plazo.get("variation"))}
                    write_row(ws, row_idx, DOLAR_MEP_CCL_COLS, fila, alt=(row_idx % 2 == 0))
                    row_idx += 1

        ws.freeze_panes = "A7"
        print(f"  Solapa DÓLAR creada ({row_idx - 7} filas de datos)")
    except Exception as e:
        import traceback
        print(f"  [ERROR] Falló al construir solapa DÓLAR: {e}")
        traceback.print_exc()


# ═══════════════════════════════════════════════════════════════════════════════
# ADRs ARGENTINA
# ═══════════════════════════════════════════════════════════════════════════════
ADR_URL = "https://api.alquimiaconsultora.com/output/yahoo_adrs_arg.json"

ADR_COLS = [
    {"key": "symbol",       "label": "Ticker",      "w":  8},
    {"key": "name",         "label": "Nombre",       "w": 28},
    {"key": "sector",       "label": "Sector",       "w": 20},
    {"key": "price",        "label": "Precio USD",  "w": 12, "align": "right", "fmt": "#,##0.00"},
    {"key": "change_pct",   "label": "Var %",        "w":  9, "align": "right", "fmt": "0.00%"},
    {"key": "day_high",     "label": "Máx. Día",    "w": 11, "align": "right", "fmt": "#,##0.00"},
    {"key": "day_low",      "label": "Mín. Día",    "w": 11, "align": "right", "fmt": "#,##0.00"},
    {"key": "volume",       "label": "Volumen",      "w": 14, "align": "right", "fmt": "#,##0"},
    {"key": "market_cap",   "label": "Market Cap",   "w": 16, "align": "right", "fmt": "#,##0"},
    {"key": "pe",           "label": "P/E",          "w":  8, "align": "right", "fmt": "#,##0.00"},
    {"key": "roa_pct",      "label": "ROA %",        "w":  9, "align": "right", "fmt": "0.00%"},
    {"key": "roe_pct",      "label": "ROE %",        "w":  9, "align": "right", "fmt": "0.00%"},
    {"key": "ytd_change",   "label": "YTD %",        "w":  9, "align": "right", "fmt": "0.00%"},
]

def build_adr(wb, fecha_dt, hora_dt):
    print(f"  Consultando {ADR_URL}...")
    try:
        r = requests.get(ADR_URL, timeout=30); r.raise_for_status()
        payload = r.json()
    except Exception as e:
        print(f"  [ERROR] No se pudo obtener datos de ADRs: {e}"); return

    raw = payload.get("adrs_argentina", [])
    if not raw:
        print("  [WARN] Sin datos de ADRs."); return

    rows = []
    for d in raw:
        rows.append({
            "symbol":     d.get("symbol", ""),
            "name":       d.get("name", ""),
            "sector":     d.get("sector", ""),
            "price":      _f(d.get("price")),
            "change_pct": _pct_direct(d.get("change_pct")),
            "day_high":   _f(d.get("day_high")),
            "day_low":    _f(d.get("day_low")),
            "volume":     _f(d.get("volume")),
            "market_cap": _f(d.get("market_cap_usd")),
            "pe":         _f(d.get("pe")),
            "roa_pct":    _pct_direct(d.get("roa_pct")),
            "roe_pct":    _pct_direct(d.get("roe_pct")),
            "ytd_change": _pct_direct(d.get("ytd_change_pct")),
        })

    rows.sort(key=lambda r: r.get("symbol") or "")

    try:
        ws = wb.create_sheet("ADR")
    except Exception:
        ws = wb.create_sheet("ADR")
    ws.sheet_properties.tabColor = TAB_ADR

    write_meta(ws, "ADRs Argentina — Cotización NYSE", fecha_dt, hora_dt, None)
    write_hdr(ws, 5, ADR_COLS)
    for i, row in enumerate(rows):
        write_row(ws, 6 + i, ADR_COLS, row, alt=(i % 2 == 1))
    ws.freeze_panes = "A6"
    print(f"  Solapa ADR creada ({len(rows)} registros)")


# ═══════════════════════════════════════════════════════════════════════════════
# NYSE / US Large Caps
# ═══════════════════════════════════════════════════════════════════════════════
NYSE_URL = "https://api.alquimiaconsultora.com/output/yahoo_nyse.json"

NYSE_COLS = [
    {"key": "symbol",       "label": "Ticker",      "w":  8},
    {"key": "name",         "label": "Nombre",       "w": 38},
    {"key": "sector",       "label": "Sector",       "w": 22},
    {"key": "price",        "label": "Precio USD",  "w": 12, "align": "right", "fmt": "#,##0.00"},
    {"key": "change_pct",   "label": "Var %",        "w":  9, "align": "right", "fmt": "0.00%"},
    {"key": "day_high",     "label": "Máx. Día",    "w": 11, "align": "right", "fmt": "#,##0.00"},
    {"key": "day_low",      "label": "Mín. Día",    "w": 11, "align": "right", "fmt": "#,##0.00"},
    {"key": "volume",       "label": "Volumen",      "w": 16, "align": "right", "fmt": "#,##0"},
    {"key": "market_cap",   "label": "Market Cap",   "w": 18, "align": "right", "fmt": "#,##0"},
    {"key": "pe",           "label": "P/E",          "w":  8, "align": "right", "fmt": "#,##0.00"},
    {"key": "roa_pct",      "label": "ROA %",        "w":  9, "align": "right", "fmt": "0.00%"},
    {"key": "roe_pct",      "label": "ROE %",        "w":  9, "align": "right", "fmt": "0.00%"},
    {"key": "ytd_change",   "label": "YTD %",        "w":  9, "align": "right", "fmt": "0.00%"},
]

def build_nyse(wb, fecha_dt, hora_dt):
    print(f"  Consultando {NYSE_URL}...")
    try:
        r = requests.get(NYSE_URL, timeout=30); r.raise_for_status()
        payload = r.json()
    except Exception as e:
        print(f"  [ERROR] No se pudo obtener datos de NYSE: {e}"); return

    raw = payload.get("nyse", [])
    if not raw:
        print("  [WARN] Sin datos de NYSE."); return

    rows = []
    for d in raw:
        rows.append({
            "symbol":     d.get("symbol", ""),
            "name":       d.get("long_name") or d.get("name", ""),
            "sector":     d.get("sector", ""),
            "price":      _f(d.get("price")),
            "change_pct": _pct_direct(d.get("change_pct")),
            "day_high":   _f(d.get("day_high")),
            "day_low":    _f(d.get("day_low")),
            "volume":     _f(d.get("volume")),
            "market_cap": _f(d.get("market_cap_usd")),
            "pe":         _f(d.get("pe")),
            "roa_pct":    _pct_direct(d.get("roa_pct")),
            "roe_pct":    _pct_direct(d.get("roe_pct")),
            "ytd_change": _pct_direct(d.get("ytd_change_pct")),
        })

    rows.sort(key=lambda r: r.get("symbol") or "")

    try:
        ws = wb.create_sheet("NYSE")
    except Exception:
        ws = wb.create_sheet("NYSE")
    ws.sheet_properties.tabColor = TAB_NYSE

    write_meta(ws, "NYSE / US Large Caps — Yahoo Finance", fecha_dt, hora_dt, None)
    write_hdr(ws, 5, NYSE_COLS)
    for i, row in enumerate(rows):
        write_row(ws, 6 + i, NYSE_COLS, row, alt=(i % 2 == 1))
    ws.freeze_panes = "A6"
    print(f"  Solapa NYSE creada ({len(rows)} registros)")


# ═══════════════════════════════════════════════════════════════════════════════
# CEDEARs / PUSH PRICES
# ═══════════════════════════════════════════════════════════════════════════════
CEDEAR_URL = "https://api.alquimiaconsultora.com/output/push_prices.json"

CEDEAR_COLS = [
    {"key": "ticker",     "label": "Ticker",      "w": 12},
    {"key": "price",      "label": "Precio",      "w": 12, "align": "right", "fmt": "#,##0.000"},
    {"key": "pct_change", "label": "Var %",       "w":  9, "align": "right", "fmt": "0.00%"},
    {"key": "q_op",       "label": "Cant. Op.",   "w": 10, "align": "right", "fmt": "#,##0"},
    {"key": "volumen",    "label": "Volumen",      "w": 18, "align": "right", "fmt": "#,##0.00"},
    {"key": "fuente",     "label": "Fuente",       "w":  8},
    {"key": "hora",       "label": "Hora",         "w":  8},
]

def build_cedear(wb, fecha_dt, hora_dt):
    print(f"  Consultando {CEDEAR_URL}...")
    try:
        r = requests.get(CEDEAR_URL, timeout=60); r.raise_for_status()
        raw = r.json()
    except Exception as e:
        print(f"  [ERROR] No se pudo obtener datos de CEDEARs: {e}"); return

    if not isinstance(raw, list):
        print("  [WARN] Formato inesperado en push_prices."); return

    # Filtrar solo registros con volumen > 0
    rows = []
    for d in raw:
        v = _f(d.get("v"))
        if not v or v <= 0: continue
        rows.append({
            "ticker":     d.get("ticker", ""),
            "price":      _f(d.get("price")),
            "pct_change": _pct_direct(d.get("pct_change")),
            "q_op":       int(d.get("q_op") or 0),
            "volumen":    v,
            "fuente":     d.get("fuente", ""),
            "hora":       d.get("hora_input", ""),
        })

    rows.sort(key=lambda r: r.get("ticker") or "")

    try:
        ws = wb.create_sheet("CEDEAR")
    except Exception:
        ws = wb.create_sheet("CEDEAR")
    ws.sheet_properties.tabColor = TAB_CEDEAR

    write_meta(ws, "CEDEARs / Push Prices — Mercado Local", fecha_dt, hora_dt, None)
    write_hdr(ws, 5, CEDEAR_COLS)
    for i, row in enumerate(rows):
        write_row(ws, 6 + i, CEDEAR_COLS, row, alt=(i % 2 == 1))
    ws.freeze_panes = "A6"
    print(f"  Solapa CEDEAR creada ({len(rows)} registros con volumen)")


# ═══════════════════════════════════════════════════════════════════════════════
# UPLOAD
# ═══════════════════════════════════════════════════════════════════════════════
def _get_auth_headers() -> dict:
    """Genera headers de autenticación usando EXPORT_SECRET."""
    secret = os.getenv("EXPORT_SECRET", "alquimia-export-2026")
    return {"X-Export-Secret": secret}


def drive_upload(fpath: Path, endpoint: str) -> str | None:
    headers = _get_auth_headers()
    try:
        with open(fpath,"rb") as f:
            resp = requests.post(
                f"{API_BASE}{endpoint}",
                headers=headers,
                files={"file": (fpath.name, f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                timeout=60)
        if resp.status_code == 200:
            data = resp.json()
            action = "Actualizado" if data.get("updated") else "Creado"
            print(f"  [DRIVE] {action} → {data.get('url') or data.get('fileId')}")
            return data.get("fileId")
        elif resp.status_code == 404:
            print(f"  [DRIVE] {endpoint} no disponible (404) — pendiente de implementar en backend")
            return None
        else:
            print(f"  [DRIVE] Error HTTP {resp.status_code}: {resp.text[:200]}")
            return None
    except Exception as e:
        print(f"  [DRIVE] Error: {e}"); return None



# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    global API_BASE, OUTPUT_DIR
    ap = argparse.ArgumentParser()
    ap.add_argument("--api",     default=API_BASE)
    ap.add_argument("--output",  default=str(OUTPUT_DIR))
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    API_BASE   = args.api.rstrip("/")
    OUTPUT_DIR = Path(args.output)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    now       = datetime.now(TZ_AR)
    fecha_hoy = now.strftime("%Y-%m-%d")
    fecha_dt  = now.date()
    hora_dt   = now.time()
    fecha_str = now.strftime("%d/%m/%Y")
    hora_str  = now.strftime("%H:%M")

    print(f"\n{'='*62}")
    print(f"  EXPORT CIERRE  v3.0  |  {fecha_str} {hora_str} ART")
    print(f"  API:     {API_BASE}")
    print(f"  OUTPUT:  {OUTPUT_DIR}")
    print(f"{'='*62}")

    if args.dry_run:
        for path in ["/api/lecaps","/api/boncer","/api/dlk","/api/tamar",
                     "/api/duales","/api/onln","/api/onlny","/api/bonoshd","/api/bote"]:
            try:
                r = requests.get(f"{API_BASE}{path}", timeout=10); d = r.json()
                rows = d.get("rows",d) if isinstance(d,dict) else d
                print(f"  ✓  {path:<18} HTTP {r.status_code}  |  {len(rows) if isinstance(rows,list) else 0} items")
            except Exception as e:
                print(f"  ✗  {path:<18} {e}")
        return

    # Cargar historia
    ultimo = OUTPUT_DIR / "ultimo_cierre.xlsx"
    print(f"\n[HIST] Cargando desde {ultimo}...")
    hist = load_historia_from_xlsx(ultimo)
    total_entradas = sum(len(v) for v in hist.values())
    print(f"  {total_entradas} entradas previas ({len(hist)} instrumentos)")
    for instr, entries in sorted(hist.items()):
        fechas = sorted(set(e.get("fecha","") for e in entries))
        print(f"    {instr}: {len(entries)} snapshots | fechas: {', '.join(fechas[-5:])}")

    # Construir workbook
    wb = Workbook(); wb.remove(wb.active)

    print("\n[1/9] LECAPS");       build_lecaps(wb, hist, fecha_hoy, hora_str, fecha_dt, hora_dt)
    print("\n[2/9] BONCER");       build_boncer(wb, hist, fecha_hoy, hora_str, fecha_dt, hora_dt)
    print("\n[3/9] DLK");          build_dlk(wb, hist, fecha_hoy, hora_str, fecha_dt, hora_dt)
    print("\n[4/9] TAMAR");        build_tamar(wb, hist, fecha_hoy, hora_str, fecha_dt, hora_dt)
    print("\n[5/9] DUALES");       build_duales(wb, hist, fecha_hoy, hora_str, fecha_dt, hora_dt)
    print("\n[6/9] ONs LN");       build_ons(wb, hist, "/api/onln",  "ONS_LN",  "ONs_LN",
                                             "ONs Ley Nacional — USD MEP / CCL",  TAB_ONLN,
                                             fecha_hoy, hora_str, fecha_dt, hora_dt)
    print("\n[7/9] ONs LNY");      build_ons(wb, hist, "/api/onlny", "ONS_LNY", "ONs_LNY",
                                             "ONs Ley Nueva York — USD MEP / CCL", TAB_ONLNY,
                                             fecha_hoy, hora_str, fecha_dt, hora_dt)
    print("\n[8/9] HD Soberanos"); build_hdsob(wb, hist, fecha_hoy, hora_str, fecha_dt, hora_dt)
    print("\n[9/9] BOTE");          build_bote(wb, hist, fecha_hoy, hora_str, fecha_dt, hora_dt)
    print("\n[+] HISTORIA");        build_historia(wb, hist)
    print("\n[+] DÓLAR");           build_dolar(wb, fecha_dt, hora_dt)
    print("\n[+] ADR");             build_adr(wb, fecha_dt, hora_dt)
    print("\n[+] NYSE");             build_nyse(wb, fecha_dt, hora_dt)
    print("\n[+] CEDEAR");          build_cedear(wb, fecha_dt, hora_dt)

    if not wb.sheetnames:
        print("\n[ERROR] Sin datos."); sys.exit(1)

    # Guardar
    fname = f"cierre_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    fpath = OUTPUT_DIR / fname
    wb.save(str(fpath))
    shutil.copy(str(fpath), str(ultimo))

    print(f"\n{'='*62}")
    print(f"  ✅  {fpath.name}")
    print(f"  ✅  ultimo_cierre.xlsx")

    print("\n[DRIVE] Subiendo ultimo_cierre.xlsx...")
    drive_upload(ultimo, "/api/admin/upload-cierre")

    print("[DRIVE] Subiendo backup diario a diarios/...")
    drive_upload(fpath, "/api/admin/upload-cierre-diario")

    print(f"\n{'='*62}\n")


if __name__ == "__main__":
    main()